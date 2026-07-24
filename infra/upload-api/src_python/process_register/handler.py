"""
ProcessRegisterFunction — Marked Register OCR Lambda

Triggered by SQS (ProcessQueue). Dispatches on message shape:

  {jobId, bucket, s3Key} where the persisted job fileType is "csv" or "xlsx"
      Parses a recognised tabular schema directly, writes one normalised JSON
      output, and settles the job as a single idempotent chunk. Spreadsheet
      inputs never invoke Poppler, Tesseract, or the PDF splitter.

  {jobId, bucket, s3Key}                                   → Splitter role.
      Counts pages, writes a JOB_CHUNKS# tracker, and enqueues one chunk
      message per page range back onto ProcessQueue. (If CHUNK_PAGES=0 it
      instead runs the original serial path unchanged — the rollback switch.)

  {jobId, bucket, s3Key, pageStart, pageEnd, chunkIndex,   → Worker role.
   totalChunks}
      OCRs that page range only (parallel across pages within the chunk),
      writes a per-chunk JSON output, and finalises the job once every chunk
      of the job has settled.

The splitter message shape is identical to what uploadCompleteHandler.mjs and
ScanResultHandlerFunction already enqueue, so those handlers need no changes.
"""

import concurrent.futures
import csv
import json
import math
import os
import re
import sys
import tempfile
import logging
import time
import uuid
import zipfile
from decimal import Decimal
from datetime import date, datetime, timezone
from pathlib import Path, PurePosixPath

import boto3
from botocore.exceptions import ClientError

# ── Tesseract / pdf2image setup ───────────────────────────────────────────────
# The Lambda layer places binaries at /opt/bin and tessdata at /opt/tessdata.
os.environ.setdefault("TESSDATA_PREFIX", "/opt/tessdata")
# Tesseract uses OpenMP internally. With several Tesseract processes running
# concurrently (OCR_WORKERS threads), each spawning its own OpenMP pool would
# oversubscribe the vCPUs and thrash — often slower than serial. Pin to 1.
os.environ.setdefault("OMP_THREAD_LIMIT", "1")
# XLSX parsing is imported lazily so PDF and CSV processing remain independent
# of the optional workbook dependency. When it is imported, force openpyxl onto
# defusedxml and away from any incidental lxml installation in a future layer.
os.environ.setdefault("OPENPYXL_DEFUSEDXML", "True")
os.environ.setdefault("OPENPYXL_LXML", "False")
POPPLER_PATH = "/opt/bin"
TESSERACT_CMD = "/opt/bin/tesseract"

try:
    import pytesseract
    from pdf2image import convert_from_path
    from pdf2image.pdf2image import pdfinfo_from_path
    from PIL import Image, ImageOps, ImageEnhance
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Config ────────────────────────────────────────────────────────────────────
JOBS_TABLE = os.environ.get("JOBS_TABLE", "")
UPLOADS_BUCKET = os.environ.get("UPLOADS_BUCKET", "")
COMBINE_FUNCTION_ARN = os.environ.get("COMBINE_FUNCTION_ARN", "")
PROCESS_QUEUE_URL = os.environ.get("PROCESS_QUEUE_URL", "")
REGION = os.environ.get("AWS_REGION", "eu-west-2")

dynamo = boto3.resource("dynamodb", region_name=REGION)
s3_client = boto3.client("s3", region_name=REGION)
lambda_client = boto3.client("lambda", region_name=REGION)
sqs_client = boto3.client("sqs", region_name=REGION)

logger = logging.getLogger()
logger.setLevel(logging.INFO)


# ════════════════════════════════════════════════════════════════════════════════
# OCR logic — ported from marked_register_processor.py
# ════════════════════════════════════════════════════════════════════════════════

_MARK_CHARS = set("—–-_.~=<>+")
_SKIP_KEYWORDS = {
    "staffordshire", "station", "register", "parliamentary", "page",
    "election", "polling", "a-1045", "cont", "index", "district council",
    "station no", "street", "general", "county",
}
_STRIKETHROUGH_SKIP = {
    "manchester", "station", "register", "parliamentary", "page",
    "election", "polling", "district", "county", "cont", "road",
    "street", "avenue", "lane", "drive",
}
_MONTH_MAP = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}
MAX_GAP_TO_FILL = 10
EVIDENCE_ONLY_GAP_INFERENCE_FLAG = "OCR_EVIDENCE_ONLY_GAP_INFERENCE"
ROW_ELIGIBILITY_FILTER_FLAG = "OCR_ROW_ELIGIBILITY_FILTER"
# Keep this aligned with the upload API and portal's accepted file-size
# contract. The row cap below remains the tighter processing safeguard for
# unusually dense CSVs.
CSV_MAX_BYTES = int(os.environ.get("CSV_MAX_BYTES", str(200 * 1024 * 1024)))
CSV_MAX_ROWS = int(os.environ.get("CSV_MAX_ROWS", "250000"))
XLSX_MAX_ARCHIVE_MEMBERS = int(
    os.environ.get("XLSX_MAX_ARCHIVE_MEMBERS", "1000")
)
XLSX_MAX_UNCOMPRESSED_BYTES = int(
    os.environ.get("XLSX_MAX_UNCOMPRESSED_BYTES", str(512 * 1024 * 1024))
)
XLSX_MAX_COMPRESSION_RATIO = int(
    os.environ.get("XLSX_MAX_COMPRESSION_RATIO", "200")
)
XLSX_MAX_WORKSHEETS = int(os.environ.get("XLSX_MAX_WORKSHEETS", "20"))
XLSX_MAX_COLUMNS = int(os.environ.get("XLSX_MAX_COLUMNS", "256"))
XLSX_MAX_CELL_CHARS = int(os.environ.get("XLSX_MAX_CELL_CHARS", "32767"))
XLSX_MAX_PHYSICAL_ROWS = int(
    os.environ.get("XLSX_MAX_PHYSICAL_ROWS", str((CSV_MAX_ROWS * 4) + 20))
)
COMBINER_CLAIM_STALE_SECONDS = int(
    os.environ.get("COMBINER_CLAIM_STALE_SECONDS", "900")
)

_AV_LIST_REQUIRED_HEADERS = {
    "polling_district": "districtref",
    "elector_number": "electorshortnumber",
    "postal_vote": "markerpostal",
}
_CSV_TRUE = frozenset({"true"})
_CSV_FALSE = frozenset({"false"})
_MARKED_POSTAL_REPORT_TITLE = "absent voter postal list marked"
_MARKED_POSTAL_REPORT_HEADERS = {
    0: "reg. no",
    2: (
        "electors name and register address "
        "postal address (if different)"
    ),
    4: "ward",
}
_PV_MARKED_REGISTER_HEADERS = frozenset({
    "recno",
    "electiondescription",
    "electiondate",
    "timefrom",
    "timeto",
    "pollingname",
    "pollingaddress1",
    "pollingaddress2",
    "pollingaddress3",
    "pollingaddress4",
    "pollingaddress5",
    "electorno",
    "electorname",
    "electoraddress1",
    "electoraddress2",
    "electoraddress3",
    "electoraddress4",
    "electoraddress5",
    "electoraddress6",
    "postalname",
    "postaladdress1",
    "postaladdress2",
    "postaladdress3",
    "postaladdress4",
    "postaladdress5",
    "postaladdress6",
    "postaladdresspostcode",
    "areaname1",
    "pvsstatus",
    "decreceiptdate",
})

_INFERENCE_DIAGNOSTIC_KEYS = (
    "numeric_gap_rows_legacy_would_generate",
    "explicit_strikethrough_rows_inferred",
    "excluded_eligibility_rows_seen",
    "excluded_eligibility_y_suppressed",
    "removed_elector_rows_excluded",
    "unreadable_strikethrough_rows_suppressed",
)

# These are the single-letter elector markers seen in UK registers. Whether a
# marker excludes an elector depends on the election: for example, B is valid at
# a local election but not a UK Parliamentary election. A is different: it marks
# an elector who may not vote in person (postal elector / postal proxy), so it is
# excluded from in-person polling-station marks in every supported register.
_KNOWN_ELIGIBILITY_CODES = frozenset("ABEFGLNP")
_ALWAYS_NOT_IN_PERSON_CODES = frozenset({"A"})
_LOCAL_ELECTION_EXCLUDED_CODES = frozenset({"E", "F"})
_PARLIAMENTARY_ELECTION_EXCLUDED_CODES = frozenset({"B", "E", "G", "L"})

# Header patterns that identify a page's polling district. Shared by the page-1
# metadata parse (_extract_metadata) and the per-page detection (§6.2).
_DISTRICT_PATTERNS = [
    r"Polling\s+District\s+([A-Z0-9]{2,8})",
    r"(?:Council|Borough|Hamlets)\s*-\s*([A-Z0-9]{2,8})\b",
    r"Electors\s+([A-Z0-9]{2,8})-\d+\s+to",
    r"\(([A-Z0-9]{2,8})-\d+\s*/\s*[A-Z0-9]+-\d+\)",
]

# Printed register declarations appear on both the cover ("Electors NAA-1 to
# NAA-926") and content-page headers ("(NAA-1 / NAA-926)"). Keep this parse
# separate from elector-row extraction: these are small, printed header bands,
# not voter rows or handwritten marks.
_DECLARED_CODE = r"[A-Z0-9]{2,8}"
_DECLARED_NUMBER = r"(?:\d{1,3}(?:,\d{3})+|\d{1,7})"
_DECLARED_RANGE_PATTERNS = [
    rf"\bElectors?\s*:?\s*({_DECLARED_CODE})\s*[-–—]\s*({_DECLARED_NUMBER})"
    rf"\s+to\s+({_DECLARED_CODE})\s*[-–—]\s*({_DECLARED_NUMBER})",
    rf"\(?\s*({_DECLARED_CODE})\s*[-–—]\s*({_DECLARED_NUMBER})"
    rf"\s*/\s*({_DECLARED_CODE})\s*[-–—]\s*({_DECLARED_NUMBER})\s*\)?",
]


def _extract_declared_ranges(text):
    """Return unique printed district/range declarations found in OCR text.

    A declaration is accepted only when the district code is printed at both
    ends and agrees case-insensitively. This intentionally rejects looser number
    pairs that commonly occur elsewhere on a register page.
    """
    ranges = []
    seen = set()
    for pattern in _DECLARED_RANGE_PATTERNS:
        for match in re.finditer(pattern, text or "", re.IGNORECASE):
            start_code, start_text, end_code, end_text = match.groups()
            district = start_code.upper()
            if end_code.upper() != district:
                continue
            try:
                start = int(start_text.replace(",", ""))
                end = int(end_text.replace(",", ""))
            except (TypeError, ValueError):
                continue
            if start < 1 or end < start:
                continue
            key = (district, start, end)
            if key in seen:
                continue
            seen.add(key)
            ranges.append({"district": district, "start": start, "end": end})
    return ranges


def _extract_polling_district_from_text(text, declared_ranges=None):
    """Preserve existing district patterns, then use a declaration as fallback."""
    for pattern in _DISTRICT_PATTERNS:
        match = re.search(pattern, text or "", re.IGNORECASE)
        if match:
            return match.group(1)
    declared_ranges = (
        _extract_declared_ranges(text) if declared_ranges is None else declared_ranges
    )
    return declared_ranges[0]["district"] if declared_ranges else None


def _row_eligibility_filter_enabled():
    return (
        os.environ.get(ROW_ELIGIBILITY_FILTER_FLAG, "false").strip().lower()
        == "true"
    )


def _extract_cover_row_rules_from_text(text):
    """Derive election-specific in-person exclusions from cover/legend OCR.

    The rules deliberately combine explicit cover wording with conservative
    election-family fallbacks. Covers vary between authorities: some print one
    comma-separated "NOT entitled" sentence, while others describe each code on
    a separate line. The fallback meanings are the standard distinctions
    demonstrated by the supplied local/county and Parliamentary examples.
    """
    text = text or ""
    lowered = re.sub(r"\s+", " ", text).lower()

    if re.search(r"\bcounty\s+(?:council\s+)?election\b", lowered):
        election_family = "local"
    elif re.search(
        r"\b(?:local government|district council|borough council|pfcc)\s+election\b",
        lowered,
    ):
        election_family = "local"
    elif re.search(
        r"\b(?:uk\s+)?parliamentary\b|\bgeneral\s+election\b",
        lowered,
    ):
        election_family = "parliamentary"
    else:
        election_family = "unknown"

    excluded_codes = set(_ALWAYS_NOT_IN_PERSON_CODES)
    if election_family == "local":
        excluded_codes.update(_LOCAL_ELECTION_EXCLUDED_CODES)
    elif election_family == "parliamentary":
        excluded_codes.update(_PARLIAMENTARY_ELECTION_EXCLUDED_CODES)

    # Examples include "F, E, printed before a name indicates ... NOT
    # entitled". Restrict the capture to isolated, comma-separated single
    # letters so prose capitals cannot accidentally become eligibility codes.
    explicit_pattern = re.compile(
        r"((?:\b[A-Z]\b\s*[,;/]\s*)+\b[A-Z]\b)\s*,?\s*"
        r"printed\s+before.{0,220}?\bNOT\s+entitled\b",
        re.IGNORECASE | re.DOTALL,
    )
    for match in explicit_pattern.finditer(text):
        excluded_codes.update(
            code.upper()
            for code in re.findall(r"\b([A-Z])\b", match.group(1), re.IGNORECASE)
            if code.upper() in _KNOWN_ELIGIBILITY_CODES
        )

    return {
        "election_family": election_family,
        "excluded_in_person_codes": sorted(excluded_codes),
    }


def _extract_cover_row_rules(image):
    """OCR a cover once and return non-personal row-classification rules."""
    text = pytesseract.image_to_string(image)
    return _extract_cover_row_rules_from_text(text)


def _extract_row_eligibility_code(line):
    """Return an isolated eligibility letter immediately after an elector number."""
    match = re.match(
        r"^[\s:;|'\"\-._~*°©=/\[\]!]*"
        r"(?:[rtli1|]\s+)?"
        r"\d+(?:\s*/\s*\d+)?\s*[:.]?\s*"
        r"[—–\-_.~=<>+]*\s*"
        r"([ABEFGLNP])(?=\s|[|&:;,./—–\-])",
        line or "",
    )
    return match.group(1).upper() if match else None


def _is_removed_elector_line(line):
    """Recognise rows explicitly labelled as removed from the register."""
    normalised = re.sub(r"[^a-z]+", " ", (line or "").lower())
    return bool(
        re.search(
            r"\belector\s+(?:has\s+been\s+)?(?:removed|rernoved)\b",
            normalised,
        )
    )


def _apply_row_eligibility_rules(line, voted, excluded_codes):
    """Return (voted, code, reason) after applying high-confidence row rules."""
    if _is_removed_elector_line(line):
        return False, None, "removed"

    code = _extract_row_eligibility_code(line)
    excluded = {
        value.upper()
        for value in (excluded_codes or _ALWAYS_NOT_IN_PERSON_CODES)
    }
    if code in excluded:
        return False, code, "excluded_eligibility"
    return voted, code, None


def _has_voting_mark(line_text, dash_chars=""):
    if dash_chars:
        mc = sum(1 for c in dash_chars if c in _MARK_CHARS)
        if mc >= 2 or any(c in dash_chars for c in "—–"):
            return True

    patterns = [
        r"[—–]", r"[._][—–\-]", r"[—–\-][._]", r"[._\-]{3,}",
        r"[<>=+][—–\-]", r"[—–\-][<>=+~]", r"~{2,}", r"={2,}", r"_{2,}",
        r"\._[\-—–]", r"[\-—–]_\.", r"^\s*[\-._]{2,}\s",
    ]
    check = line_text[:35]
    if any(re.search(p, check) for p in patterns):
        return True

    early = re.match(r"^\d+(?:/\d+)?\s*([\-]{2,})", line_text)
    if early:
        return True

    total = len(re.findall(r"[—–\-_~]", line_text))
    if total >= 4:
        h = len(line_text) // 2
        if (
            len(re.findall(r"[—–\-_~]", line_text[:h])) >= 1
            and len(re.findall(r"[—–\-_~]", line_text[h:])) >= 1
        ):
            return True

    if re.search(r"[A-Z][a-z]+[\-—–][A-Z][a-z]+[A-Z]", line_text):
        return True
    if re.search(r"[A-Z][a-z]+[\-—–][A-Z][a-z]*[éèêëàâäùûü]", line_text):
        return True
    if re.search(r"[A-Z][a-z]+\s+[A-Z][a-z]+[\-—–][\s]*$", line_text):
        return True
    return False


def _extract_elector_entry(line, context_prev_num=0):
    line = line.strip()
    if not line or len(line) < 5:
        return None, None

    line = re.sub(r"^[\s:;|'\"\-._~*°©=/\[\]!]+", "", line)

    noise = re.match(r"^([rtl1|i])\s+(\d{2,3})", line)
    if noise:
        line = line[noise.end(1):].strip()

    if re.match(r"^[tT]\s+[A-Z]", line) and context_prev_num in range(4, 7):
        line = "7" + line[1:]
    if re.match(r"^(ai|TT|tt|//)\s+", line) and 75 <= context_prev_num <= 78:
        line = "77" + line[2:]
    if line.startswith("2295"):
        line = "225" + line[4:]

    m227 = re.match(r"^22\s*/\s*([—–\-])", line)
    if m227 and 225 <= context_prev_num <= 228:
        line = "227 " + m227.group(1) + line[m227.end():]

    merged = re.match(r"^(\d{2,3})([4-9])\s*[\-—–]+", line)
    if merged and context_prev_num > 0:
        p = int(merged.group(1))
        if context_prev_num < p <= context_prev_num + 3:
            line = merged.group(1) + " ——" + line[merged.end():]

    if context_prev_num == 180 and line.startswith("184"):
        line = "181" + line[3:]

    line_lower = line.lower()
    if any(kw in line_lower for kw in _SKIP_KEYWORDS):
        return None, None
    if re.match(r"^[A-Za-z][a-z]+\s*$", line):
        return None, None

    m = re.match(
        r"^[}\[\]|:.\s/]*(\d+(?:\s*/\s*\d+)?)\s*[:.]?\s*([—–\-_.~=<>+]*)\s*(.*)$",
        line,
    )
    if not m:
        return None, None

    elector_num = re.sub(r"\s*/\s*", "/", m.group(1))
    dash_chars = m.group(2) or ""
    rest = m.group(3) or ""

    if rest and re.match(r"^\d+\s*$", rest):
        return None, None
    if elector_num.startswith("0"):
        return None, None

    parts = elector_num.split("/")
    try:
        main_num = int(parts[0])
        sub_num = int(parts[1]) if len(parts) > 1 else 0
    except ValueError:
        return None, None

    if context_prev_num > 0:
        if main_num < 10 and 70 <= context_prev_num < 80:
            c = 70 + main_num
            if context_prev_num < c <= context_prev_num + 5:
                main_num = c
                elector_num = f"{main_num}/{sub_num}" if sub_num else str(main_num)
        elif 10 <= main_num < 20 and 70 <= context_prev_num < 80:
            c = 60 + main_num
            if context_prev_num < c <= context_prev_num + 5:
                main_num = c
                elector_num = f"{main_num}/{sub_num}" if sub_num else str(main_num)
        elif main_num < 20 and context_prev_num >= 20:
            decade = (context_prev_num // 10) * 10
            c = decade + main_num
            if context_prev_num < c <= context_prev_num + 15:
                main_num = c
                elector_num = f"{main_num}/{sub_num}" if sub_num else str(main_num)

    if main_num > 2000 or main_num < 1:
        return None, None
    if not re.search(r"[A-Za-z]{2,}", rest):
        return None, None
    if rest and re.match(r"^[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s*$", rest.strip()):
        if "," not in rest and len(rest.strip().split()) <= 2:
            return None, None

    voted = _has_voting_mark(line, dash_chars)
    if not voted and rest:
        em = re.match(r"^[—–\-_.~=<>+]{1,}", rest[:10])
        if em and (any(c in em.group() for c in "—–") or len(em.group()) >= 2):
            voted = True

    return elector_num, voted


def _is_likely_strikethrough(line):
    ll = line.lower()
    if any(kw in ll for kw in _STRIKETHROUGH_SKIP) or len(line) < 10:
        return False
    garbled = len(re.findall(r"[A-Z]{3,}[a-z]|[a-z][A-Z]{2,}|[mwnr]{3,}", line))
    mixed = len(re.findall(r"[a-z][A-Z][a-z]|[A-Z][a-z][A-Z]", line))
    repeated = len(re.findall(r"(.)\1{2,}", line))
    dashes = len(re.findall(r"[—–\-_=]{2,}", line))
    has_name = bool(re.search(r"[A-Z][a-z]{2,}", line))
    num_let = len(re.findall(r"\d[A-Za-z]|[A-Za-z]\d", line))
    score = garbled + mixed + repeated + dashes + num_let
    return (score >= 3 and (has_name or dashes >= 1)) or (len(line) > 40 and garbled >= 2)


def _new_inference_diagnostics():
    return {key: 0 for key in _INFERENCE_DIAGNOSTIC_KEYS}


def _merge_inference_diagnostics(target, source):
    for key in _INFERENCE_DIAGNOSTIC_KEYS:
        target[key] += int((source or {}).get(key, 0))


def _evidence_only_gap_inference_enabled():
    return os.environ.get(EVIDENCE_ONLY_GAP_INFERENCE_FLAG, "false").strip().lower() == "true"


def _infer_missing_entries(readable_entries, start_num,
                           evidence_only_gap_inference=None, diagnostics=None,
                           row_eligibility_filter=None):
    if not readable_entries:
        return []
    if evidence_only_gap_inference is None:
        evidence_only_gap_inference = _evidence_only_gap_inference_enabled()
    if row_eligibility_filter is None:
        row_eligibility_filter = _row_eligibility_filter_enabled()
    entries = []
    expected = start_num + 1 if start_num > 0 else 1
    # A strikethrough with no readable number can only be inferred once we have an
    # anchor to count from — either a starting context (start_num > 0) or a
    # readable number already seen in this column. Without one, `expected` is just
    # the default 1, so inferring would fabricate elector "1" (and, with the
    # district resolver, a spurious district) on every column. Skip until anchored.
    anchored = start_num > 0
    for entry in readable_entries:
        if entry.get("is_strikethrough") and entry["main_num"] is None:
            if not anchored:
                continue
            if row_eligibility_filter:
                if diagnostics is not None:
                    diagnostics["unreadable_strikethrough_rows_suppressed"] += 1
                # A numberless struck-through row cannot reveal whether the line
                # is an attendant mark, a printed eligibility exclusion, or ink
                # bleeding from an adjacent row. Do not manufacture a Y.
                continue
            entries.append({"elector_num": str(expected), "voted": True})
            if diagnostics is not None:
                diagnostics["explicit_strikethrough_rows_inferred"] += 1
            expected += 1
        elif entry["main_num"] is not None:
            actual = entry["main_num"]
            if not anchored:
                # First readable number in the column (and no starting context):
                # anchor to it. There is no basis to fill any gap that precedes the
                # first readable number, so a leading low reading like 11 must not
                # fabricate 1..10.
                expected = actual
            gap = actual - expected
            if 0 < gap <= MAX_GAP_TO_FILL:
                if diagnostics is not None:
                    diagnostics["numeric_gap_rows_legacy_would_generate"] += gap
                if not evidence_only_gap_inference:
                    while expected < actual:
                        entries.append({"elector_num": str(expected), "voted": True})
                        expected += 1
            elif gap > MAX_GAP_TO_FILL:
                expected = actual
            entries.append({"elector_num": entry["elector_num"], "voted": entry["voted"]})
            expected = actual + 1
            anchored = True
    return entries


def _process_column(image, col_start, col_end, context_start_num=0,
                    inference_diagnostics=None, row_eligibility_filter=None,
                    excluded_codes=None):
    if row_eligibility_filter is None:
        row_eligibility_filter = _row_eligibility_filter_enabled()
    excluded_codes = set(excluded_codes or _ALWAYS_NOT_IN_PERSON_CODES)
    col_img = image.crop((col_start, 0, col_end, image.size[1]))
    col_img = ImageOps.expand(col_img, border=70, fill="white")
    col_img = ImageEnhance.Contrast(col_img).enhance(1.3)

    text = pytesseract.image_to_string(
        col_img, config=r"--oem 3 --psm 6 -c preserve_interword_spaces=1"
    )

    readable = []
    prev_num = context_start_num
    for line in text.split("\n"):
        line = line.strip()
        if not line or len(line) < 3:
            continue
        if row_eligibility_filter and _is_removed_elector_line(line):
            if inference_diagnostics is not None:
                inference_diagnostics["removed_elector_rows_excluded"] += 1
            continue
        elector_num, voted = _extract_elector_entry(line, prev_num)
        if elector_num:
            if row_eligibility_filter:
                raw_voted = voted
                voted, eligibility_code, reason = _apply_row_eligibility_rules(
                    line, voted, excluded_codes
                )
                if reason == "removed":
                    # Kept as a defensive branch for OCR variants that only
                    # become recognisable after elector parsing.
                    continue
                if reason == "excluded_eligibility":
                    if inference_diagnostics is not None:
                        inference_diagnostics["excluded_eligibility_rows_seen"] += 1
                        if raw_voted:
                            inference_diagnostics[
                                "excluded_eligibility_y_suppressed"
                            ] += 1
            try:
                main_num = int(elector_num.split("/")[0])
                if context_start_num > 10 and main_num < context_start_num - 10:
                    continue
                readable.append(
                    {"elector_num": elector_num, "main_num": main_num, "voted": voted, "line": line}
                )
                prev_num = main_num
            except ValueError:
                pass
        elif _is_likely_strikethrough(line):
            readable.append(
                {"elector_num": None, "main_num": None, "voted": True, "line": line, "is_strikethrough": True}
            )

    entries = _infer_missing_entries(
        readable, context_start_num, diagnostics=inference_diagnostics,
        row_eligibility_filter=row_eligibility_filter,
    )

    last_num = context_start_num
    if entries:
        try:
            last_num = max(int(e["elector_num"].split("/")[0]) for e in entries if e.get("elector_num"))
        except (ValueError, AttributeError):
            pass
    return entries, last_num


def _detect_columns(image):
    width, height = image.size
    mid = image.crop((int(width * 0.33), 0, int(width * 0.66), height))
    mid = ImageOps.expand(mid, border=70, fill="white")
    mid = ImageEnhance.Contrast(mid).enhance(1.3)
    text = pytesseract.image_to_string(mid, config=r"--oem 3 --psm 6")
    count = sum(
        1 for line in text.split("\n")
        if re.match(r"^\d{1,4}\s*[—–\-]?\s*[A-Z][a-z]+", line.strip())
    )
    return 3 if count >= 10 else 2


def _process_page(image, context_num=0, ncols=None, inference_diagnostics=None,
                  row_eligibility_filter=None, excluded_codes=None):
    w = image.size[0]
    # Column layout is a property of the register's print format and is fixed for
    # the whole document, so the caller may pass a cached ncols (detected once per
    # chunk in §5.6) to skip the extra full-page Tesseract pass. When None, detect
    # as before — this preserves the original serial behaviour exactly.
    if ncols is None:
        ncols = _detect_columns(image)
    entries = []
    if ncols == 3:
        for start, end in [(0, int(w * 0.32)), (int(w * 0.32), int(w * 0.64)), (int(w * 0.64), w)]:
            col_entries, _ = _process_column(
                image, start, end, 0, inference_diagnostics,
                row_eligibility_filter, excluded_codes,
            )
            entries.extend(col_entries)
        last_num = context_num
    else:
        left, left_last = _process_column(
            image, 0, int(w * 0.48), 0, inference_diagnostics,
            row_eligibility_filter, excluded_codes,
        )
        right, right_last = _process_column(
            image, int(w * 0.50), w, 0, inference_diagnostics,
            row_eligibility_filter, excluded_codes,
        )
        entries = left + right
        last_num = max(left_last, right_last) if left_last and right_last else (left_last or right_last or 0)
    return entries, last_num


def _extract_metadata(image):
    """Extract date, polling district, vote type, and declared ranges."""
    text = pytesseract.image_to_string(image)

    # Election date
    dm = re.search(
        r"(\d{1,2})\s*(January|February|March|April|May|June|July|August|September|"
        r"October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})",
        text, re.IGNORECASE,
    )
    election_date = None
    if dm:
        month = _MONTH_MAP.get(dm.group(2).lower(), "01")
        election_date = f"{dm.group(1).zfill(2)}/{month}/{dm.group(3)}"

    vote_type = "Postal" if "postal" in text.lower() else "In Person"
    declared_ranges = _extract_declared_ranges(text)
    # Existing patterns remain authoritative for byte-identical behaviour on
    # registers that already work. A parsed declaration is the fallback ground
    # truth when spacing or a typographic dash defeats those older patterns.
    polling_district = _extract_polling_district_from_text(text, declared_ranges)
    return election_date, polling_district or "Unknown", vote_type, declared_ranges


def _extract_page_header(image):
    """OCR a page header and return (district code, declared ranges).

    Reads only the top ~12% of the page (printed header text, never marks), so it
    is cheap and safe to run once per page inside the worker loop (§6.2)."""
    w, h = image.size
    header = image.crop((0, 0, w, int(h * 0.12)))
    text = pytesseract.image_to_string(header, config=r"--oem 3 --psm 6")
    declared_ranges = _extract_declared_ranges(text)
    district = _extract_polling_district_from_text(text, declared_ranges)
    return district, declared_ranges


def _extract_page_district(image):
    """Backward-compatible district-only wrapper used by pure callers/tests."""
    district, _ = _extract_page_header(image)
    return district


def _render_page(pdf_path, page_num, grayscale=False):
    """Render a single page at 600dpi. Greyscale is opt-in (OCR_GRAYSCALE) and
    cuts image memory ~3x; Tesseract binarises internally so it must be verified
    byte-identical (Test 1) before the default is flipped (§5.8)."""
    return convert_from_path(
        pdf_path, dpi=600, first_page=page_num, last_page=page_num,
        poppler_path=POPPLER_PATH, grayscale=grayscale,
    )


def _count_pages(pdf_path):
    try:
        info = pdfinfo_from_path(pdf_path, poppler_path=POPPLER_PATH)
        return int(info.get("Pages", 0))
    except Exception:
        return 0


def _build_rows(entries, constituency_name, election_date, polling_district, vote_type,
                attach_page):
    """Map extracted entries to CSV row dicts. attach_page adds the source page
    number (consumed by combiner district resolution; build_csv ignores it)."""
    rows = []
    for e in entries:
        row = {
            "election_date": election_date,
            "constituency": constituency_name or "Unknown Constituency",
            "polling_district": polling_district,
            "elector_number": e["elector_num"],
            "voted": "Y" if e.get("voted") else "N",
            "postal_vote": "Y" if vote_type == "Postal" else "N",
        }
        if attach_page and e.get("page") is not None:
            row["page"] = e["page"]
        rows.append(row)
    return rows


class CsvInputError(ValueError):
    """A permanent, privacy-safe tabular-input validation failure."""

    def __init__(self, code, message):
        super().__init__(f"{code}: {message}")
        self.code = code
        self.detail = message


class XlsxInputError(CsvInputError):
    """A permanent, privacy-safe XLSX validation failure."""


def _normalise_csv_header(value):
    """Normalise a heading for exact alias matching, never fuzzy matching."""
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().casefold())


def _detect_csv_delimiter(header_line):
    """Choose a delimiter from the header only.

    Header-only detection avoids feeding names, addresses, or other row data to
    a sniffer while still accepting the common CSV/TSV exports produced by
    electoral-management systems.
    """
    counts = {delimiter: header_line.count(delimiter) for delimiter in ",;\t|"}
    highest = max(counts.values(), default=0)
    winners = [delimiter for delimiter, count in counts.items() if count == highest]
    if highest <= 0 or len(winners) != 1:
        raise CsvInputError(
            "CSV_DELIMITER_INVALID",
            "Could not identify one supported delimiter from the header row.",
        )
    return winners[0]


def _resolve_av_list_columns(headers):
    """Return source-column indexes for the recognised absent-voter schema."""
    by_normalised = {}
    duplicate_heading = False
    for index, header in enumerate(headers):
        normalised = _normalise_csv_header(header)
        if not normalised:
            continue
        if normalised in by_normalised:
            duplicate_heading = True
        else:
            by_normalised[normalised] = index
    if duplicate_heading:
        raise CsvInputError(
            "CSV_HEADER_INVALID",
            "The CSV contains duplicate headings after normalisation.",
        )

    missing = [
        source_name
        for source_name in _AV_LIST_REQUIRED_HEADERS.values()
        if source_name not in by_normalised
    ]
    if missing:
        raise CsvInputError(
            "CSV_HEADER_UNRECOGNISED",
            "The CSV does not match a supported marked-register input schema.",
        )
    return {
        target: by_normalised[source_name]
        for target, source_name in _AV_LIST_REQUIRED_HEADERS.items()
    }


def _parse_absent_voter_marker(value):
    normalised = str(value or "").strip().casefold()
    if normalised in _CSV_TRUE:
        return "Y"
    if normalised in _CSV_FALSE:
        return "N"
    return None


def _normalise_report_cell(value):
    """Collapse presentation whitespace while retaining punctuation."""
    return " ".join(str(value or "").split()).strip().casefold()


def _nonempty_columns(row):
    return tuple(
        index for index, value in enumerate(row)
        if str(value or "").strip()
    )


def _is_marked_postal_report_title(row):
    return (
        len(row) == 9
        and _nonempty_columns(row) == (0,)
        and _normalise_report_cell(row[0]) == _MARKED_POSTAL_REPORT_TITLE
    )


def _is_marked_postal_report_header(row):
    if len(row) != 9 or _nonempty_columns(row) != (0, 2, 4):
        return False
    return all(
        _normalise_report_cell(row[index]) == expected
        for index, expected in _MARKED_POSTAL_REPORT_HEADERS.items()
    )


def _is_pv_marked_register_header(row):
    """Recognise the exact thirty-column postal-vote marked-register export."""
    normalised = [
        _normalise_csv_header(value)
        for value in row
    ]
    return (
        len(normalised) == len(_PV_MARKED_REGISTER_HEADERS)
        and len(set(normalised)) == len(normalised)
        and set(normalised) == _PV_MARKED_REGISTER_HEADERS
    )


def _valid_pv_receipt_date(value):
    """Accept the export's timestamp and Excel's date-only representation."""
    text = str(value or "").strip()
    if not text:
        return False
    for date_format in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            datetime.strptime(text, date_format)
            return True
        except ValueError:
            continue
    return False


def _parse_pv_marked_register_rows(
    reader,
    headers,
    constituency,
    election_date,
    source_format,
):
    """Normalise the thirty-column postal-vote marked-register export.

    Only ElectorNo, PVSStatus, and DecReceiptDate are inspected. Names,
    addresses, polling-place details, and other presentation columns are
    structurally bounded by the shared XLSX reader but are never retained.
    """
    normalised_headers = [
        _normalise_csv_header(value)
        for value in headers
    ]
    indexes = {
        header: normalised_headers.index(header)
        for header in ("electorno", "pvsstatus", "decreceiptdate")
    }
    expected_width = len(headers)
    rows = []
    seen_keys = set()
    invalid_width = 0
    invalid_references = 0
    invalid_statuses = 0
    invalid_receipt_dates = 0
    inconsistent_markers = 0
    duplicate_keys = 0
    data_rows_seen = 0
    voted_rows = 0

    for source_row in reader:
        if not source_row or all(
            not str(value or "").strip() for value in source_row
        ):
            continue
        data_rows_seen += 1
        if data_rows_seen > CSV_MAX_ROWS:
            raise CsvInputError(
                "CSV_TOO_MANY_ROWS",
                "The CSV exceeds the configured row limit.",
            )
        if len(source_row) != expected_width:
            invalid_width += 1
            continue

        reference = str(source_row[indexes["electorno"]] or "").strip()
        status = str(source_row[indexes["pvsstatus"]] or "").strip()
        receipt_date = str(
            source_row[indexes["decreceiptdate"]] or ""
        ).strip()

        match = re.fullmatch(
            r"([A-Za-z0-9]{2,8})-([1-9]\d*(?:/\d+)?)",
            reference,
        )
        status_valid = re.fullmatch(r"\d{1,10}", status) is not None
        receipt_present = bool(receipt_date)
        receipt_valid = (
            not receipt_present
            or _valid_pv_receipt_date(receipt_date)
        )
        row_invalid = False
        if not match:
            invalid_references += 1
            row_invalid = True
        if not status_valid:
            invalid_statuses += 1
            row_invalid = True
        if not receipt_valid:
            invalid_receipt_dates += 1
            row_invalid = True

        status_marked = (
            status_valid
            and any(character != "0" for character in status)
        )
        if (
            status_valid
            and receipt_valid
            and status_marked != receipt_present
        ):
            inconsistent_markers += 1
            row_invalid = True
        if row_invalid:
            continue

        key = (match.group(1).upper(), match.group(2))
        if key in seen_keys:
            duplicate_keys += 1
            continue
        seen_keys.add(key)
        if receipt_present:
            voted_rows += 1
        rows.append({
            "election_date": election_date,
            "constituency": constituency,
            "polling_district": key[0],
            "elector_number": key[1],
            "voted": "Y" if receipt_present else "N",
            "postal_vote": "Y",
        })

    validation_counts = {
        "wrong-width rows": invalid_width,
        "rows with invalid ElectorNo": invalid_references,
        "rows with invalid PVSStatus": invalid_statuses,
        "rows with invalid DecReceiptDate": invalid_receipt_dates,
        "rows with inconsistent postal-vote markers": inconsistent_markers,
        "duplicate district/elector keys": duplicate_keys,
    }
    failures = [
        f"{label}: {count}"
        for label, count in validation_counts.items()
        if count
    ]
    if failures:
        raise CsvInputError(
            "CSV_ROW_INVALID",
            "The CSV failed aggregate row validation ("
            + "; ".join(failures)
            + ").",
        )
    if not rows:
        raise CsvInputError(
            "CSV_EMPTY",
            "The CSV file is empty or contains no data rows.",
        )

    return rows, {
        "source_type": "csv",
        "source_format": source_format,
        "csv_schema": "pv_marked_register_v1",
        "rows_read": len(rows),
        "postal_rows": len(rows),
        "voted_rows": voted_rows,
        "declared_ranges": [],
        "inference_diagnostics": _new_inference_diagnostics(),
    }


def _next_report_row(reader):
    try:
        return next(reader)
    except StopIteration as exc:
        raise CsvInputError(
            "CSV_REPORT_INVALID",
            "The marked postal-list report has an incomplete record structure.",
        ) from exc


def _parse_marked_postal_report(
    reader,
    first_row,
    constituency,
    election_date,
    source_format="csv",
):
    """Stream the nine-column multi-record marked postal-list report.

    Each elector is represented by one register-reference row followed by a
    separate name/address detail row. Only the register reference is parsed;
    the sensitive detail row is checked structurally and then discarded.
    """
    expected_width = len(first_row)

    date_row = _next_report_row(reader)
    metadata_row = _next_report_row(reader)
    pre_header_blank = _next_report_row(reader)
    header_row = _next_report_row(reader)
    if (
        len(date_row) != expected_width
        or _nonempty_columns(date_row) != (5,)
        or len(metadata_row) != expected_width
        or _nonempty_columns(metadata_row) != (0,)
        or len(pre_header_blank) != expected_width
        or _nonempty_columns(pre_header_blank)
        or not _is_marked_postal_report_header(header_row)
    ):
        raise CsvInputError(
            "CSV_REPORT_INVALID",
            "The marked postal-list report preamble or headings are invalid.",
        )

    # The report has two blank records between its headings and first block.
    for _ in range(2):
        spacer = _next_report_row(reader)
        if (
            len(spacer) != expected_width
            or _nonempty_columns(spacer)
        ):
            raise CsvInputError(
                "CSV_REPORT_INVALID",
                "The marked postal-list report has an invalid record structure.",
            )

    rows = []
    seen_keys = set()
    invalid_references = 0
    duplicate_keys = 0
    candidate_blocks = 0
    state = "candidate"
    logical_records = 7

    for source_row in reader:
        logical_records += 1
        if logical_records > (CSV_MAX_ROWS * 4) + 20:
            raise CsvInputError(
                "CSV_TOO_MANY_ROWS",
                "The CSV exceeds the configured row limit.",
            )
        if len(source_row) != expected_width:
            raise CsvInputError(
                "CSV_REPORT_INVALID",
                "The marked postal-list report has an invalid record width.",
            )

        signature = _nonempty_columns(source_row)
        if state == "candidate":
            if not signature:
                if candidate_blocks == 0:
                    raise CsvInputError(
                        "CSV_REPORT_INVALID",
                        "The marked postal-list report contains no data blocks.",
                    )
                state = "footer_label"
                continue
            if signature != (0, 5):
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report has an invalid data-block structure.",
                )

            candidate_blocks += 1
            match = re.fullmatch(
                r"([A-Za-z0-9]{2,8})-([1-9]\d*(?:/\d+)?)",
                source_row[0].strip(),
            )
            if not match:
                invalid_references += 1
            else:
                key = (match.group(1).upper(), match.group(2))
                if key in seen_keys:
                    duplicate_keys += 1
                else:
                    seen_keys.add(key)
                    rows.append({
                        "election_date": election_date,
                        "constituency": constituency,
                        "polling_district": key[0],
                        "elector_number": key[1],
                        # The report title is not row-level evidence that a
                        # ballot was returned.
                        "voted": "N",
                        # This profile is explicitly a postal-list report.
                        "postal_vote": "Y",
                    })
                    if len(rows) > CSV_MAX_ROWS:
                        raise CsvInputError(
                            "CSV_TOO_MANY_ROWS",
                            "The CSV exceeds the configured row limit.",
                        )
            state = "block_blank"
        elif state == "block_blank":
            if signature:
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report has an invalid data-block structure.",
                )
            state = "private_detail"
        elif state == "private_detail":
            if signature != (0,):
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report has an invalid data-block structure.",
                )
            # Deliberately do not inspect or retain the name/address cell.
            state = "block_end"
        elif state == "block_end":
            if signature:
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report has an invalid data-block structure.",
                )
            state = "candidate"
        elif state == "footer_label":
            total_match = re.fullmatch(
                r"total number of postal voters (\d+)",
                _normalise_report_cell(source_row[0])
                if signature == (0,)
                else "",
            )
            if (
                not total_match
                or int(total_match.group(1)) != candidate_blocks
            ):
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report total does not match its data blocks.",
                )
            state = "footer_blank"
        elif state == "footer_blank":
            if signature:
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report footer is invalid.",
                )
            state = "footer_total"
        elif state == "footer_total":
            generated_date = (
                _normalise_report_cell(source_row[0])
                if signature == (0, 8)
                else ""
            )
            page_match = re.fullmatch(
                r"page (\d+)-of-(\d+)",
                _normalise_report_cell(source_row[8])
                if signature == (0, 8)
                else "",
            )
            if (
                not re.fullmatch(r"\d{2}/\d{2}/\d{4}", generated_date)
                or not page_match
                or page_match.group(1) != page_match.group(2)
            ):
                raise CsvInputError(
                    "CSV_REPORT_INVALID",
                    "The marked postal-list report footer is invalid.",
                )
            state = "done"
        else:
            raise CsvInputError(
                "CSV_REPORT_INVALID",
                "The marked postal-list report has trailing records.",
            )

    if state != "done":
        raise CsvInputError(
            "CSV_REPORT_INVALID",
            "The marked postal-list report has an incomplete record structure.",
        )
    if invalid_references or duplicate_keys:
        failures = []
        if invalid_references:
            failures.append(
                f"rows with invalid register references: {invalid_references}"
            )
        if duplicate_keys:
            failures.append(
                f"duplicate district/elector keys: {duplicate_keys}"
            )
        raise CsvInputError(
            "CSV_ROW_INVALID",
            "The CSV failed aggregate row validation ("
            + "; ".join(failures)
            + ").",
        )
    if not rows:
        raise CsvInputError(
            "CSV_EMPTY",
            "The CSV file is empty or contains no data rows.",
        )

    return rows, {
        "source_type": "csv",
        "source_format": source_format,
        "csv_schema": "marked_postal_report_v1",
        "rows_read": len(rows),
        "postal_rows": len(rows),
        "declared_ranges": [],
        "inference_diagnostics": _new_inference_diagnostics(),
    }


def _parse_flat_absent_voter_rows(
    reader,
    headers,
    constituency,
    election_date,
    source_format,
):
    """Normalise the flat absent-voter profile from a string-row iterator."""
    indexes = _resolve_av_list_columns(headers)
    expected_width = len(headers)
    rows = []
    seen_keys = set()
    invalid_width = 0
    missing_district = 0
    invalid_district = 0
    missing_elector = 0
    invalid_elector = 0
    invalid_postal = 0
    duplicate_keys = 0
    data_rows_seen = 0

    for source_row in reader:
        if not source_row or all(
            not str(value or "").strip() for value in source_row
        ):
            continue
        data_rows_seen += 1
        if data_rows_seen > CSV_MAX_ROWS:
            raise CsvInputError(
                "CSV_TOO_MANY_ROWS",
                "The CSV exceeds the configured row limit.",
            )
        if len(source_row) != expected_width:
            invalid_width += 1
            continue

        district = source_row[indexes["polling_district"]].strip().upper()
        elector_number = source_row[indexes["elector_number"]].strip()
        postal_vote = _parse_absent_voter_marker(
            source_row[indexes["postal_vote"]]
        )

        row_invalid = False
        if not district:
            missing_district += 1
            row_invalid = True
        elif not re.fullmatch(r"[A-Z0-9]{2,8}", district):
            invalid_district += 1
            row_invalid = True
        if not elector_number:
            missing_elector += 1
            row_invalid = True
        elif not re.fullmatch(r"[1-9]\d*(?:/\d+)?", elector_number):
            invalid_elector += 1
            row_invalid = True
        if postal_vote is None:
            invalid_postal += 1
            row_invalid = True
        if row_invalid:
            continue

        key = (district, elector_number)
        if key in seen_keys:
            duplicate_keys += 1
            continue
        seen_keys.add(key)
        rows.append({
            "election_date": election_date,
            "constituency": constituency,
            "polling_district": district,
            "elector_number": elector_number,
            # An absent-vote arrangement is not evidence that a ballot was
            # returned. A matching PDF row may promote this to Y in the
            # combiner.
            "voted": "N",
            "postal_vote": postal_vote,
        })

    validation_counts = {
        "wrong-width rows": invalid_width,
        "rows missing DistrictRef": missing_district,
        "rows with invalid DistrictRef": invalid_district,
        "rows missing ElectorShortNumber": missing_elector,
        "rows with invalid ElectorShortNumber": invalid_elector,
        "rows with invalid MarkerPostal": invalid_postal,
        "duplicate district/elector keys": duplicate_keys,
    }
    failures = [f"{label}: {count}" for label, count in validation_counts.items() if count]
    if failures:
        raise CsvInputError(
            "CSV_ROW_INVALID",
            "The CSV failed aggregate row validation (" + "; ".join(failures) + ").",
        )
    if not rows:
        raise CsvInputError(
            "CSV_EMPTY",
            "The CSV file is empty or contains no data rows.",
        )

    postal_rows = sum(1 for row in rows if row["postal_vote"] == "Y")
    meta = {
        "source_type": "csv",
        "source_format": source_format,
        "csv_schema": "absent_voters_v1",
        "rows_read": len(rows),
        "postal_rows": postal_rows,
        "declared_ranges": [],
        "inference_diagnostics": _new_inference_diagnostics(),
    }
    return rows, meta


def _parse_tabular_rows(
    reader,
    headers,
    constituency,
    election_date,
    *,
    source_format,
    report_allowed=True,
):
    """Dispatch a canonical string-row iterator to one supported schema."""
    if _is_pv_marked_register_header(headers):
        return _parse_pv_marked_register_rows(
            reader,
            headers,
            constituency,
            election_date,
            source_format,
        )
    if _is_marked_postal_report_title(headers):
        if not report_allowed:
            raise CsvInputError(
                "CSV_HEADER_UNRECOGNISED",
                "The CSV does not match a supported marked-register input schema.",
            )
        return _parse_marked_postal_report(
            reader,
            headers,
            constituency,
            election_date,
            source_format=source_format,
        )
    return _parse_flat_absent_voter_rows(
        reader,
        headers,
        constituency,
        election_date,
        source_format,
    )


def _parse_uploaded_csv(csv_path, constituency_name, election_date):
    """Parse one recognised CSV into the same internal row shape as PDF OCR.

    Recognised profiles are the flat absent-voter export
    (DistrictRef + ElectorShortNumber + MarkerPostal) and the nine-column
    multi-record marked postal-list report, plus the thirty-column PV marked
    register export (ElectorNo + PVSStatus + DecReceiptDate). Unrelated or
    sensitive presentation columns are used only for structural framing and are
    never copied to JSON, logs, or errors.
    """
    constituency = str(constituency_name or "").strip()
    election_date = str(election_date or "").strip()
    if not constituency or not election_date:
        raise CsvInputError(
            "CSV_METADATA_MISSING",
            "CSV processing requires constituency and election-date upload metadata.",
        )

    size = os.path.getsize(csv_path)
    if size <= 0:
        raise CsvInputError(
            "CSV_EMPTY",
            "The CSV file is empty or contains no data rows.",
        )
    if size > CSV_MAX_BYTES:
        raise CsvInputError(
            "CSV_TOO_LARGE",
            "The CSV exceeds the configured processing size limit.",
        )

    try:
        with open(csv_path, "r", encoding="utf-8-sig", newline="") as handle:
            header_line = handle.readline()
            if not header_line.strip():
                raise CsvInputError(
                    "CSV_EMPTY",
                    "The CSV file is empty or contains no data rows.",
                )
            delimiter = _detect_csv_delimiter(header_line)
            handle.seek(0)
            reader = csv.reader(handle, delimiter=delimiter, strict=True)
            try:
                headers = next(reader)
            except StopIteration as exc:
                raise CsvInputError(
                    "CSV_EMPTY",
                    "The CSV file is empty or contains no data rows.",
                ) from exc
            return _parse_tabular_rows(
                reader,
                headers,
                constituency,
                election_date,
                source_format="csv",
                report_allowed=(delimiter == ","),
            )
    except UnicodeDecodeError as exc:
        raise CsvInputError(
            "CSV_ENCODING_INVALID",
            "The CSV must use UTF-8 encoding.",
        ) from exc
    except csv.Error as exc:
        raise CsvInputError(
            "CSV_PARSE_INVALID",
            "The CSV is not valid RFC-4180-style delimited text.",
        ) from exc


def _translate_xlsx_validation_error(exc):
    code = (
        f"XLSX_{exc.code[4:]}"
        if exc.code.startswith("CSV_")
        else exc.code
    )
    detail = (
        str(exc.detail)
        .replace("The CSV", "The Excel workbook")
        .replace("CSV processing", "XLSX processing")
    )
    return XlsxInputError(code, detail)


def _validate_xlsx_archive(xlsx_path):
    """Reject unsafe or implausibly large OOXML containers before XML parsing."""
    size = os.path.getsize(xlsx_path)
    if size <= 0:
        raise XlsxInputError(
            "XLSX_EMPTY",
            "The Excel workbook is empty or contains no data rows.",
        )
    if size > CSV_MAX_BYTES:
        raise XlsxInputError(
            "XLSX_TOO_LARGE",
            "The Excel workbook exceeds the configured processing size limit.",
        )

    required_members = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
    }
    allowed_compression = {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}
    total_compressed = 0
    total_uncompressed = 0
    try:
        with zipfile.ZipFile(xlsx_path, "r", allowZip64=True) as archive:
            members = archive.infolist()
            if (
                not members
                or len(members) > XLSX_MAX_ARCHIVE_MEMBERS
            ):
                raise XlsxInputError(
                    "XLSX_ARCHIVE_INVALID",
                    "The Excel workbook archive has an invalid structure.",
                )

            names = {member.filename for member in members}
            if not required_members.issubset(names):
                raise XlsxInputError(
                    "XLSX_ARCHIVE_INVALID",
                    "The upload is not a supported XLSX workbook.",
                )

            for member in members:
                raw_name = member.filename
                normalised_name = raw_name.replace("\\", "/")
                member_path = PurePosixPath(normalised_name)
                lower_name = normalised_name.casefold()
                if (
                    raw_name != normalised_name
                    or member_path.is_absolute()
                    or ".." in member_path.parts
                    or member.flag_bits & 0x1
                    or member.compress_type not in allowed_compression
                ):
                    raise XlsxInputError(
                        "XLSX_ARCHIVE_INVALID",
                        "The Excel workbook archive has an invalid structure.",
                    )
                if (
                    lower_name == "xl/vbaproject.bin"
                    or lower_name.startswith("xl/activex/")
                    or lower_name.startswith("xl/embeddings/")
                ):
                    raise XlsxInputError(
                        "XLSX_ACTIVE_CONTENT_UNSUPPORTED",
                        "The Excel workbook contains unsupported active content.",
                    )
                if member.file_size > XLSX_MAX_UNCOMPRESSED_BYTES:
                    raise XlsxInputError(
                        "XLSX_ARCHIVE_TOO_LARGE",
                        "The expanded Excel workbook exceeds the configured limit.",
                    )
                total_compressed += member.compress_size
                total_uncompressed += member.file_size
                if total_uncompressed > XLSX_MAX_UNCOMPRESSED_BYTES:
                    raise XlsxInputError(
                        "XLSX_ARCHIVE_TOO_LARGE",
                        "The expanded Excel workbook exceeds the configured limit.",
                    )

            if (
                total_uncompressed > 0
                and (
                    total_compressed <= 0
                    or total_uncompressed
                    > total_compressed * XLSX_MAX_COMPRESSION_RATIO
                )
            ):
                raise XlsxInputError(
                    "XLSX_ARCHIVE_RATIO_INVALID",
                    "The Excel workbook expansion ratio exceeds the configured limit.",
                )
    except (zipfile.BadZipFile, zipfile.LargeZipFile, NotImplementedError) as exc:
        raise XlsxInputError(
            "XLSX_ARCHIVE_INVALID",
            "The upload is not a valid XLSX workbook.",
        ) from exc


def _xlsx_cell_to_text(cell, *, reject_unsafe=True):
    """Convert one openpyxl cell without evaluating formulas."""
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        if reject_unsafe:
            raise XlsxInputError(
                "XLSX_FORMULA_UNSUPPORTED",
                "Formula cells are not supported in uploaded workbooks.",
            )
        return ""
    if data_type == "e":
        if reject_unsafe:
            raise XlsxInputError(
                "XLSX_CELL_INVALID",
                "The Excel workbook contains an invalid cell value.",
            )
        return ""

    value = getattr(cell, "value", None)
    if value is None:
        text = ""
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            raise XlsxInputError(
                "XLSX_CELL_INVALID",
                "The Excel workbook contains an invalid numeric cell.",
            )
        text = str(int(value)) if value.is_integer() else format(value, ".15g")
    elif isinstance(value, datetime):
        text = value.strftime("%d/%m/%Y")
    elif isinstance(value, date):
        text = value.strftime("%d/%m/%Y")
    elif isinstance(value, str):
        text = value
    else:
        raise XlsxInputError(
            "XLSX_CELL_INVALID",
            "The Excel workbook contains an unsupported cell value.",
        )

    if len(text) > XLSX_MAX_CELL_CHARS:
        raise XlsxInputError(
            "XLSX_CELL_TOO_LARGE",
            "The Excel workbook contains a cell that exceeds the configured limit.",
        )
    return text


def _trim_trailing_blank_cells(values):
    end = len(values)
    while end and not str(values[end - 1] or "").strip():
        end -= 1
    return values[:end]


def _xlsx_header_candidate(worksheet):
    """Inspect row one only and return a schema name without logging values."""
    scan_columns = min(
        max(int(worksheet.max_column or 1), 1),
        257,
    )
    cells = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            min_col=1,
            max_col=scan_columns,
        ),
        (),
    )
    values = _trim_trailing_blank_cells([
        _xlsx_cell_to_text(cell, reject_unsafe=False)
        for cell in cells
    ])
    if (
        len(values) <= 9
        and _is_marked_postal_report_title(values + [""] * (9 - len(values)))
    ):
        return "report"
    if _is_pv_marked_register_header(values):
        return "pv_marked"

    normalised = {
        _normalise_csv_header(value)
        for value in values
        if str(value or "").strip()
    }
    if set(_AV_LIST_REQUIRED_HEADERS.values()).issubset(normalised):
        return "flat"
    return None


def _xlsx_data_rows(worksheet, expected_width):
    """Yield canonical rows, padding sparse cells and flagging overflow safely."""
    max_column = max(int(worksheet.max_column or 1), 1)
    for physical_row, cells in enumerate(
        worksheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=max_column,
        ),
        start=2,
    ):
        if physical_row > XLSX_MAX_PHYSICAL_ROWS:
            raise XlsxInputError(
                "XLSX_TOO_MANY_ROWS",
                "The Excel workbook exceeds the configured row limit.",
            )
        values = [_xlsx_cell_to_text(cell) for cell in cells]
        overflow = any(
            str(value or "").strip()
            for value in values[expected_width:]
        )
        canonical = values[:expected_width]
        if len(canonical) < expected_width:
            canonical.extend([""] * (expected_width - len(canonical)))
        if overflow:
            # The shared validator counts this as a wrong-width row without
            # retaining or echoing the out-of-range cell value.
            canonical.append("")
        yield canonical


def _parse_uploaded_xlsx(xlsx_path, constituency_name, election_date):
    """Read exactly one recognised worksheet through the shared normaliser."""
    constituency = str(constituency_name or "").strip()
    election_date = str(election_date or "").strip()
    if not constituency or not election_date:
        raise XlsxInputError(
            "XLSX_METADATA_MISSING",
            "XLSX processing requires constituency and election-date upload metadata.",
        )

    _validate_xlsx_archive(xlsx_path)

    try:
        import openpyxl
        from defusedxml.common import DefusedXmlException
        from openpyxl.utils.exceptions import InvalidFileException
        from xml.etree.ElementTree import ParseError
    except ImportError as exc:
        raise RuntimeError(
            "Secure XLSX parser dependencies are unavailable."
        ) from exc
    if not getattr(openpyxl, "DEFUSEDXML", False):
        raise RuntimeError("Secure XLSX XML parsing is not enabled.")

    malformed_errors = (
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        InvalidFileException,
        ParseError,
        DefusedXmlException,
        KeyError,
        TypeError,
        ValueError,
    )
    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            xlsx_path,
            read_only=True,
            data_only=False,
            keep_links=False,
            keep_vba=False,
            rich_text=False,
        )
        worksheets = list(workbook.worksheets)
        if (
            not worksheets
            or len(worksheets) > XLSX_MAX_WORKSHEETS
        ):
            raise XlsxInputError(
                "XLSX_WORKBOOK_INVALID",
                "The Excel workbook has an invalid worksheet structure.",
            )

        candidates = [
            (worksheet, schema)
            for worksheet in worksheets
            if (schema := _xlsx_header_candidate(worksheet)) is not None
        ]
        if not candidates:
            raise XlsxInputError(
                "XLSX_HEADER_UNRECOGNISED",
                "The Excel workbook does not contain one supported marked-register worksheet.",
            )
        if len(candidates) != 1:
            raise XlsxInputError(
                "XLSX_WORKSHEET_AMBIGUOUS",
                "The Excel workbook contains more than one supported marked-register worksheet.",
            )

        worksheet, schema = candidates[0]
        if worksheet.sheet_state != "visible":
            raise XlsxInputError(
                "XLSX_WORKSHEET_HIDDEN",
                "The supported marked-register worksheet must be visible.",
            )
        if int(worksheet.max_column or 1) > XLSX_MAX_COLUMNS:
            raise XlsxInputError(
                "XLSX_TOO_MANY_COLUMNS",
                "The Excel workbook exceeds the configured column limit.",
            )
        if int(worksheet.max_row or 1) > XLSX_MAX_PHYSICAL_ROWS:
            raise XlsxInputError(
                "XLSX_TOO_MANY_ROWS",
                "The Excel workbook exceeds the configured row limit.",
            )

        header_cells = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                min_col=1,
                max_col=max(int(worksheet.max_column or 1), 1),
            ),
            (),
        )
        headers = _trim_trailing_blank_cells([
            _xlsx_cell_to_text(cell)
            for cell in header_cells
        ])
        if schema == "report":
            if len(headers) > 9:
                raise XlsxInputError(
                    "XLSX_REPORT_INVALID",
                    "The marked postal-list report has an invalid record width.",
                )
            headers.extend([""] * (9 - len(headers)))
            expected_width = 9
        else:
            expected_width = len(headers)

        try:
            return _parse_tabular_rows(
                _xlsx_data_rows(worksheet, expected_width),
                headers,
                constituency,
                election_date,
                source_format="xlsx",
                report_allowed=True,
            )
        except XlsxInputError:
            raise
        except CsvInputError as exc:
            raise _translate_xlsx_validation_error(exc) from exc
    except XlsxInputError:
        raise
    except malformed_errors as exc:
        raise XlsxInputError(
            "XLSX_WORKBOOK_INVALID",
            "The upload is not a valid XLSX workbook.",
        ) from exc
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                logger.warning("Could not close XLSX workbook cleanly")


def _ocr_serial(pdf_path, total_pages, constituency_name, election_date,
                polling_district, vote_type, inference_diagnostics=None,
                row_eligibility_filter=None, excluded_codes=None):
    """Original serial page-by-page OCR path, preserved unchanged for the
    CHUNK_PAGES=0 rollback switch (§10)."""
    if total_pages == 0:
        # Fallback: convert all and measure (original behaviour; rollback only).
        all_pages = convert_from_path(pdf_path, dpi=600, poppler_path=POPPLER_PATH)
        total_pages = len(all_pages)

    skip_pages = 2
    all_entries = []

    for page_num in range(1, total_pages + 1):
        if page_num <= skip_pages:
            logger.info("Skipping cover page %d", page_num)
            continue

        logger.info("OCR page %d / %d", page_num, total_pages)
        page_images = convert_from_path(
            pdf_path, dpi=600,
            first_page=page_num, last_page=page_num,
            poppler_path=POPPLER_PATH,
        )
        if not page_images:
            continue

        img = page_images[0]
        entries, _ = _process_page(
            img,
            inference_diagnostics=inference_diagnostics,
            row_eligibility_filter=row_eligibility_filter,
            excluded_codes=excluded_codes,
        )
        all_entries.extend(entries)
        del img, page_images  # free memory; image file not persisted

    # Deduplicate by elector number (original whole-document key).
    seen = set()
    unique = []
    for e in all_entries:
        k = e.get("elector_num")
        if k and k not in seen:
            seen.add(k)
            unique.append(e)

    return _build_rows(unique, constituency_name, election_date, polling_district,
                       vote_type, attach_page=False)


def _ocr_chunk(pdf_path, page_start, page_end, total_pages, constituency_name,
               election_date, polling_district, vote_type,
               inference_diagnostics=None, row_eligibility_filter=None,
               excluded_codes=None):
    """Parallel OCR of one page range. Detects column layout once for the chunk,
    OCRs the remaining pages concurrently, tags every row with its source page,
    and records each page's detected polling district and declared elector range.
    Returns (rows, pageDistricts, pageDeclaredRanges)."""
    workers = int(os.environ.get("OCR_WORKERS", "6"))
    grayscale = os.environ.get("OCR_GRAYSCALE", "false").strip().lower() == "true"
    skip_pages = 2

    if total_pages > 0:
        if page_start > total_pages:
            # A chunk pointed entirely beyond the document is a hard failure, not
            # a silently-empty success (Test 7).
            raise ValueError(
                f"Chunk page range {page_start}-{page_end} is beyond document length {total_pages}"
            )
        if page_end is None or page_end > total_pages:
            page_end = total_pages

    # Absolute page numbering keeps the cover-page skip absolute (invariant 3).
    content_pages = [p for p in range(page_start, page_end + 1) if p > skip_pages]

    page_entries = {}     # absolute page number -> list of entry dicts
    page_districts = {}   # absolute page number -> district code or None
    page_declared_ranges = {}  # absolute page number -> printed range declarations

    if not content_pages:
        return [], page_districts, page_declared_ranges

    # §5.6 — detect the column layout once, on the chunk's first content page, and
    # reuse that rendered image so we don't render it a second time at 600dpi.
    detect_page = content_pages[0]
    ncols = None
    first_imgs = _render_page(pdf_path, detect_page, grayscale)
    if first_imgs:
        img0 = first_imgs[0]
        try:
            ncols = _detect_columns(img0)
            page_inference_diagnostics = _new_inference_diagnostics()
            entries0, _ = _process_page(
                img0, ncols=ncols,
                inference_diagnostics=page_inference_diagnostics,
                row_eligibility_filter=row_eligibility_filter,
                excluded_codes=excluded_codes,
            )
            if inference_diagnostics is not None:
                _merge_inference_diagnostics(
                    inference_diagnostics, page_inference_diagnostics
                )
            for e in entries0:
                e["page"] = detect_page
            page_entries[detect_page] = entries0
            district, declared_ranges = _extract_page_header(img0)
            page_districts[detect_page] = district
            page_declared_ranges[detect_page] = declared_ranges
        finally:
            del img0, first_imgs

    def _work(page_num):
        imgs = _render_page(pdf_path, page_num, grayscale)
        if not imgs:
            return page_num, [], None, [], _new_inference_diagnostics()
        img = imgs[0]
        try:
            page_inference_diagnostics = _new_inference_diagnostics()
            entries, _ = _process_page(
                img, ncols=ncols,
                inference_diagnostics=page_inference_diagnostics,
                row_eligibility_filter=row_eligibility_filter,
                excluded_codes=excluded_codes,
            )
            for e in entries:
                e["page"] = page_num
            district, declared_ranges = _extract_page_header(img)
            return (
                page_num, entries, district, declared_ranges,
                page_inference_diagnostics,
            )
        finally:
            del img, imgs

    remaining = content_pages[1:]
    if remaining:
        # Threads (not processes): pytesseract and pdf2image both shell out to
        # external binaries, releasing the GIL, and threads avoid pickling PIL
        # images / fork() issues inside Lambda (§5.2).
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
            for (page_num, entries, district, declared_ranges,
                 page_inference_diagnostics) in pool.map(_work, remaining):
                page_entries[page_num] = entries
                page_districts[page_num] = district
                page_declared_ranges[page_num] = declared_ranges
                if inference_diagnostics is not None:
                    _merge_inference_diagnostics(
                        inference_diagnostics, page_inference_diagnostics
                    )

    # Flatten in ascending page order — ordering must be deterministic (invariant 1).
    all_entries = []
    for page_num in sorted(page_entries):
        all_entries.extend(page_entries[page_num])

    # In-worker dedupe keyed on (page, elector_num). The duplicates this removes
    # are within-page OCR artefacts (the same line caught by two overlapping
    # column crops). Cross-page repeats are almost always distinct electors in
    # different districts and must survive to the combiner (§5.2 / §6).
    seen = set()
    unique = []
    for e in all_entries:
        key = (e.get("page"), e.get("elector_num"))
        if key[1] and key not in seen:
            seen.add(key)
            unique.append(e)

    rows = _build_rows(unique, constituency_name, election_date, polling_district,
                       vote_type, attach_page=True)
    # JSON object keys are strings; keep the map string-keyed for round-trip
    # stability through S3 (§5.4).
    page_districts_out = {str(k): v for k, v in page_districts.items()}
    page_declared_ranges_out = {
        str(k): v for k, v in page_declared_ranges.items()
    }
    return rows, page_districts_out, page_declared_ranges_out


def ocr_pdf(pdf_path, constituency_name, election_name, election_date_override="",
            page_start=None, page_end=None):
    """
    OCR a PDF and return (rows, meta, pageDistricts, pageDeclaredRanges).

    When page_start/page_end are given the function OCRs only that page range in
    parallel (chunk worker mode), tagging each row with its source page and
    returning per-page district and declared-range maps. When both are omitted it
    runs the original serial path unchanged (the CHUNK_PAGES=0 rollback),
    returning empty page maps and rows without a page field.

    Page-1 metadata (date, polling district, vote type) is extracted at 150dpi in
    every invocation, so all chunks of a job carry the same seed values (§6.3).
    election_date_override (from the form) takes precedence over OCR-derived dates.
    """
    logger.info("Converting first page for metadata extraction")
    first_pages = convert_from_path(
        pdf_path, dpi=150, first_page=1, last_page=1, poppler_path=POPPLER_PATH
    )
    ocr_election_date, polling_district, vote_type, declared_ranges = (
        _extract_metadata(first_pages[0])
        if first_pages else (None, "Unknown", "In Person", [])
    )
    row_eligibility_filter = _row_eligibility_filter_enabled()
    row_rules = (
        _extract_cover_row_rules(first_pages[0])
        if row_eligibility_filter and first_pages
        else {
            "election_family": "unknown",
            "excluded_in_person_codes": sorted(_ALWAYS_NOT_IN_PERSON_CODES),
        }
    )
    if row_eligibility_filter and row_rules["election_family"] == "unknown":
        form_rules = _extract_cover_row_rules_from_text(election_name)
        if form_rules["election_family"] != "unknown":
            row_rules = form_rules
    excluded_codes = row_rules["excluded_in_person_codes"]
    election_date = (
        election_date_override
        or ocr_election_date
        or datetime.now(timezone.utc).strftime("%d/%m/%Y")
    )
    meta = {
        "election_date": election_date,
        "polling_district": polling_district,
        "vote_type": vote_type,
        "declared_ranges": declared_ranges,
        "row_eligibility_rules": row_rules,
    }
    inference_diagnostics = _new_inference_diagnostics()

    total_pages = _count_pages(pdf_path)

    if page_start is None:
        if row_eligibility_filter:
            rows = _ocr_serial(
                pdf_path, total_pages, constituency_name, election_date,
                polling_district, vote_type, inference_diagnostics,
                row_eligibility_filter=True, excluded_codes=excluded_codes,
            )
        else:
            # Preserve the historical call shape for the default-off path and
            # rollback tests.
            rows = _ocr_serial(
                pdf_path, total_pages, constituency_name, election_date,
                polling_district, vote_type, inference_diagnostics,
            )
        meta["inference_diagnostics"] = inference_diagnostics
        logger.info(
            "OCR inference diagnostics: "
            "numeric_gap_rows_legacy_would_generate=%d, "
            "explicit_strikethrough_rows_inferred=%d, "
            "excluded_eligibility_rows_seen=%d, "
            "excluded_eligibility_y_suppressed=%d, "
            "removed_elector_rows_excluded=%d, "
            "unreadable_strikethrough_rows_suppressed=%d",
            inference_diagnostics["numeric_gap_rows_legacy_would_generate"],
            inference_diagnostics["explicit_strikethrough_rows_inferred"],
            inference_diagnostics["excluded_eligibility_rows_seen"],
            inference_diagnostics["excluded_eligibility_y_suppressed"],
            inference_diagnostics["removed_elector_rows_excluded"],
            inference_diagnostics["unreadable_strikethrough_rows_suppressed"],
        )
        logger.info("OCR complete (serial): %d entries extracted", len(rows))
        return rows, meta, {}, {}

    rows, page_districts, page_declared_ranges = _ocr_chunk(
        pdf_path, page_start, page_end, total_pages, constituency_name,
        election_date, polling_district, vote_type, inference_diagnostics,
        row_eligibility_filter=row_eligibility_filter,
        excluded_codes=excluded_codes,
    )
    meta["inference_diagnostics"] = inference_diagnostics
    logger.info(
        "OCR inference diagnostics: "
        "numeric_gap_rows_legacy_would_generate=%d, "
        "explicit_strikethrough_rows_inferred=%d, "
        "excluded_eligibility_rows_seen=%d, "
        "excluded_eligibility_y_suppressed=%d, "
        "removed_elector_rows_excluded=%d, "
        "unreadable_strikethrough_rows_suppressed=%d",
        inference_diagnostics["numeric_gap_rows_legacy_would_generate"],
        inference_diagnostics["explicit_strikethrough_rows_inferred"],
        inference_diagnostics["excluded_eligibility_rows_seen"],
        inference_diagnostics["excluded_eligibility_y_suppressed"],
        inference_diagnostics["removed_elector_rows_excluded"],
        inference_diagnostics["unreadable_strikethrough_rows_suppressed"],
    )
    logger.info(
        "OCR complete (chunk %s-%s): %d entries extracted", page_start, page_end, len(rows)
    )
    return rows, meta, page_districts, page_declared_ranges


# ════════════════════════════════════════════════════════════════════════════════
# DynamoDB helpers
# ════════════════════════════════════════════════════════════════════════════════

def get_job(job_id):
    table = dynamo.Table(JOBS_TABLE)
    resp = table.get_item(Key={"jobId": job_id})
    return resp.get("Item")


def update_job_succeeded(job_id, output_key, processed_at):
    table = dynamo.Table(JOBS_TABLE)
    table.update_item(
        Key={"jobId": job_id},
        UpdateExpression="SET #st = :s, outputKey = :ok, processedAt = :pa, updatedAt = :ua",
        ExpressionAttributeNames={"#st": "status"},
        ExpressionAttributeValues={
            ":s": "SUCCEEDED",
            ":ok": output_key,
            ":pa": processed_at,
            ":ua": processed_at,
        },
    )


def update_job_failed(job_id, reason, failed_at):
    table = dynamo.Table(JOBS_TABLE)
    table.update_item(
        Key={"jobId": job_id},
        UpdateExpression="SET #st = :s, failureReason = :r, updatedAt = :ua",
        ExpressionAttributeNames={"#st": "status"},
        ExpressionAttributeValues={":s": "FAILED", ":r": reason, ":ua": failed_at},
    )


class CombinerClaimPending(RuntimeError):
    """A retryable hand-off state; the current job outcome remains terminal."""


def try_trigger_combiner(
    batch_id,
    total_files,
    job_payload,
    job_id,
    *,
    count_completion=True,
):
    """
    Record one job completion idempotently and invoke the combiner when ready.

    `completedJobs` and `completedCount` are updated together under a condition
    that the current job ID is not already present. A retry after an ambiguous
    DynamoDB response therefore cannot increment the numeric counter twice.
    Legacy trackers that only have `completedCount` remain compatible.

    If this Lambda completes the batch, it claims combiner invocation via
    a conditional update. If the asynchronous invoke is rejected, the caller's
    own tokenised claim is released so an SQS redelivery can retry it.
    As with any asynchronous network hand-off, an ambiguous transport failure
    can produce at-least-once delivery; the combiner's stable S3 output key and
    terminal status writes are repeatable.

    Called once per file when the file's job settles — success OR failure — so a
    failed file never leaves the batch hanging (§5.7.1).
    """
    table = dynamo.Table(JOBS_TABLE)
    tracker_key = f"BATCH_TRACKER#{batch_id}"

    if count_completion:
        try:
            resp = table.update_item(
                Key={"jobId": tracker_key},
                UpdateExpression=(
                    "ADD completedJobs :jobs, completedCount :one "
                    "SET totalFiles = if_not_exists(totalFiles, :tf)"
                ),
                ConditionExpression=(
                    "attribute_not_exists(completedJobs) "
                    "OR NOT contains(completedJobs, :job)"
                ),
                ExpressionAttributeValues={
                    ":jobs": {job_id},
                    ":job": job_id,
                    ":one": Decimal("1"),
                    ":tf": Decimal(str(total_files)),
                },
                ReturnValues="ALL_NEW",
            )
            attrs = resp["Attributes"]
        except ClientError as exc:
            if exc.response["Error"]["Code"] != "ConditionalCheckFailedException":
                raise
            # This job was already counted. A consistent read also covers a
            # retry after the first update committed but its response was lost.
            attrs = (
                table.get_item(
                    Key={"jobId": tracker_key},
                    ConsistentRead=True,
                ).get("Item")
                or {}
            )
    else:
        # Compatibility path for a JOB_CHUNKS tracker carrying the old
        # `batchCounted` marker: do not increment its legacy count again.
        attrs = (
            table.get_item(
                Key={"jobId": tracker_key},
                ConsistentRead=True,
            ).get("Item")
            or {}
        )

    completed = int(attrs.get("completedCount", 0))
    logger.info("Batch %s: %d / %d completed", batch_id, completed, total_files)

    if completed < total_files:
        return

    # Try to claim the combiner invocation. A timestamped, unaccepted claim can
    # be stolen after one full function timeout, recovering the narrow crash
    # window between the claim write and the async invoke call.
    claim_token = f"{job_id}:{uuid.uuid4().hex}"
    claim_time = int(time.time())
    stale_before = claim_time - COMBINER_CLAIM_STALE_SECONDS
    try:
        table.update_item(
            Key={"jobId": tracker_key},
            UpdateExpression=(
                "SET combinerInvoked = :claim, combinerClaimedAt = :claimed_at "
                "REMOVE combinerAcceptedAt"
            ),
            ConditionExpression=(
                "attribute_not_exists(combinerInvoked) OR "
                "(attribute_not_exists(combinerAcceptedAt) "
                "AND combinerClaimedAt < :stale_before)"
            ),
            ExpressionAttributeValues={
                ":claim": claim_token,
                ":claimed_at": Decimal(str(claim_time)),
                ":stale_before": Decimal(str(stale_before)),
            },
        )
    except ClientError as e:
        if e.response["Error"]["Code"] == "ConditionalCheckFailedException":
            current = (
                table.get_item(
                    Key={"jobId": tracker_key},
                    ConsistentRead=True,
                ).get("Item")
                or {}
            )
            if current.get("combinerAcceptedAt") is not None:
                logger.info("Batch %s: combiner invocation already accepted", batch_id)
                return
            if current.get("combinerClaimedAt") is None:
                # Legacy boolean/token claims predate the accepted-at marker and
                # remain authoritative for in-flight deployment compatibility.
                logger.info("Batch %s: legacy combiner claim already present", batch_id)
                return
            raise CombinerClaimPending(
                "Combiner invocation is currently claimed but not yet accepted."
            ) from e
        raise

    logger.info("Batch %s: invoking CombineRegisterFunction", batch_id)
    try:
        invoke_response = lambda_client.invoke(
            FunctionName=COMBINE_FUNCTION_ARN,
            InvocationType="Event",  # async
            Payload=json.dumps(job_payload).encode(),
        )
        if int((invoke_response or {}).get("StatusCode", 0)) != 202:
            raise RuntimeError("CombineRegisterFunction did not accept the event.")
    except Exception:
        logger.exception(
            "Batch %s: combiner invoke failed — releasing invocation claim",
            batch_id,
        )
        try:
            table.update_item(
                Key={"jobId": tracker_key},
                UpdateExpression="REMOVE combinerInvoked, combinerClaimedAt",
                ConditionExpression="combinerInvoked = :claim",
                ExpressionAttributeValues={":claim": claim_token},
            )
        except ClientError as release_exc:
            if (
                release_exc.response["Error"]["Code"]
                != "ConditionalCheckFailedException"
            ):
                logger.exception(
                    "Batch %s: failed to release combiner invocation claim",
                    batch_id,
                )
        raise
    try:
        table.update_item(
            Key={"jobId": tracker_key},
            UpdateExpression="SET combinerAcceptedAt = :accepted_at",
            ConditionExpression="combinerInvoked = :claim",
            ExpressionAttributeValues={
                ":accepted_at": Decimal(str(int(time.time()))),
                ":claim": claim_token,
            },
        )
    except Exception:
        # The async service has already accepted the event. Do not fail the SQS
        # message and risk a duplicate invocation solely because this
        # observability marker could not be written.
        logger.exception(
            "Batch %s: accepted combiner event but could not mark it accepted",
            batch_id,
        )


def try_finalise_job(job_id, chunk_index, batch_id, total_files, combine_payload,
                     output_prefix, chunk_failed, failure_reason=None):
    """
    Settle one chunk of a job and, once every chunk has settled, finalise the job.

    Chunk settlement is outcome-idempotent: a conditional update records each
    chunk index in DynamoDB number sets exactly once. A redelivery reads and
    honours the first settlement rather than changing success into failure (or
    vice versa). The job is complete when the set reaches totalChunks.

    On completion the job's terminal status is written (idempotent SET), then
    the batch tracker records this job ID and increments its count in one
    conditional update. Nothing here is swallowed: transient DynamoDB errors
    propagate to the caller, which returns the message to SQS for an idempotent
    retry rather than stranding the job.

    Raises on any DynamoDB error so the caller can retry. Returns True once the
    job has been finalised (or was already), False while chunks remain.
    """
    table = dynamo.Table(JOBS_TABLE)
    tracker_key = f"JOB_CHUNKS#{job_id}"

    if chunk_failed:
        update_expr = "ADD settledChunks :chunk_set, failedChunks :chunk_set"
    else:
        update_expr = "ADD settledChunks :chunk_set"
    try:
        resp = table.update_item(
            Key={"jobId": tracker_key},
            UpdateExpression=update_expr,
            ConditionExpression=(
                "attribute_not_exists(settledChunks) "
                "OR NOT contains(settledChunks, :chunk_index)"
            ),
            ExpressionAttributeValues={
                ":chunk_set": {chunk_index},
                ":chunk_index": Decimal(str(chunk_index)),
            },
            ReturnValues="ALL_NEW",
        )
        attrs = resp["Attributes"]
    except ClientError as exc:
        if exc.response["Error"]["Code"] != "ConditionalCheckFailedException":
            raise
        # The first successful settlement owns this chunk's outcome. Read it
        # back consistently so a later max-receive redelivery cannot promote a
        # previously successful chunk into failedChunks (or vice versa).
        attrs = (
            table.get_item(
                Key={"jobId": tracker_key},
                ConsistentRead=True,
            ).get("Item")
            or {}
        )
    total_chunks = int(attrs.get("totalChunks", 0))
    settled = attrs.get("settledChunks") or set()
    failed = attrs.get("failedChunks") or set()
    logger.info(
        "Job %s: chunk %s settled (%d / %d chunks, %d failed)",
        job_id, chunk_index, len(settled), total_chunks, len(failed),
    )

    if total_chunks <= 0 or len(settled) < total_chunks:
        return False

    # Write the job's terminal status. SET to a fixed value is idempotent, so it is
    # safe to run more than once across concurrent finishers or a redelivery.
    now_iso = datetime.now(timezone.utc).isoformat()
    if failed:
        reason = failure_reason or f"{len(failed)} of {total_chunks} chunk(s) failed"
        update_job_failed(job_id, reason, now_iso)
        logger.warning("Job %s finalised as FAILED (%d/%d chunks failed)",
                       job_id, len(failed), total_chunks)
    else:
        update_job_succeeded(job_id, output_prefix, now_iso)
        logger.info("Job %s finalised as SUCCEEDED (%d chunks)", job_id, total_chunks)

    if not batch_id:
        return True

    try_trigger_combiner(
        batch_id,
        total_files,
        combine_payload,
        job_id,
        count_completion=not bool(attrs.get("batchCounted")),
    )
    return True


# ════════════════════════════════════════════════════════════════════════════════
# Splitter
# ════════════════════════════════════════════════════════════════════════════════

def _input_file_type(job, s3_key):
    """Use persisted API metadata first; suffix fallback is legacy-only."""
    value = (
        (job or {}).get("fileType")
        or (job or {}).get("expectedFileType")
        or Path(s3_key or "").suffix.lstrip(".")
    )
    normalised = str(value or "").strip().lower()
    if normalised not in {"pdf", "csv", "xlsx"}:
        raise ValueError("Unsupported input file type.")
    return normalised


def _ensure_single_chunk_tracker(job_id):
    """Initialise the tracker used by a direct CSV or XLSX job."""
    table = dynamo.Table(JOBS_TABLE)
    response = table.update_item(
        Key={"jobId": f"JOB_CHUNKS#{job_id}"},
        UpdateExpression="SET totalChunks = if_not_exists(totalChunks, :one)",
        ExpressionAttributeValues={":one": Decimal("1")},
        ReturnValues="ALL_NEW",
    )
    return response.get("Attributes") or {}


def _build_chunk_ranges(total_pages, chunk_pages):
    ranges = []
    page = 1
    while page <= total_pages:
        ranges.append((page, min(page + chunk_pages - 1, total_pages)))
        page += chunk_pages
    return ranges


def run_splitter(job_id, bucket, s3_key, chunk_pages):
    """Count pages, write the JOB_CHUNKS# tracker, and enqueue one chunk message
    per page range. The tracker is written BEFORE any chunk is enqueued so a fast
    chunk can never complete against a tracker that does not yet exist (§5.1)."""
    suffix = Path(s3_key).suffix or ".pdf"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = tmp.name
    try:
        s3_client.download_file(bucket, s3_key, tmp_path)
        try:
            info = pdfinfo_from_path(tmp_path, poppler_path=POPPLER_PATH)
        except Exception as exc:
            # Do NOT fall back to a full 600dpi render just to count pages — that
            # is exactly the silent 900s burn this change removes (§5.1).
            raise ValueError(f"Could not read PDF page count for {s3_key}: {exc}") from exc
        total_pages = int(info.get("Pages", 0))
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    if total_pages <= 0:
        raise ValueError(f"PDF {s3_key} reported {total_pages} pages")

    ranges = _build_chunk_ranges(total_pages, chunk_pages)
    total_chunks = len(ranges)

    # Write the tracker first (totalChunks known). The settledChunks/failedChunks
    # sets are created lazily by the first chunk's ADD — DynamoDB has no empty set.
    table = dynamo.Table(JOBS_TABLE)
    table.update_item(
        Key={"jobId": f"JOB_CHUNKS#{job_id}"},
        UpdateExpression="SET totalChunks = :tc",
        ExpressionAttributeValues={":tc": Decimal(str(total_chunks))},
    )

    for chunk_index, (page_start, page_end) in enumerate(ranges):
        message = {
            "jobId": job_id,
            "bucket": bucket,
            "s3Key": s3_key,
            "pageStart": page_start,
            "pageEnd": page_end,
            "chunkIndex": chunk_index,
            "totalChunks": total_chunks,
        }
        sqs_client.send_message(
            QueueUrl=PROCESS_QUEUE_URL,
            MessageBody=json.dumps(message),
        )

    logger.info(
        "Splitter: job %s (%d pages) -> %d chunk(s) of up to %d pages",
        job_id, total_pages, total_chunks, chunk_pages,
    )


# ════════════════════════════════════════════════════════════════════════════════
# Lambda handler
# ════════════════════════════════════════════════════════════════════════════════

def _cleanup(path):
    try:
        os.unlink(path)
    except OSError:
        pass


def handler(event, context):
    chunk_pages = int(os.environ.get("CHUNK_PAGES", "20"))
    records = event.get("Records", [])
    failures = []

    for record in records:
        job_id = None
        mode = "serial"
        settle_ctx = None
        try:
            body = json.loads(record["body"])
            job_id = body["jobId"]
            bucket = body.get("bucket", UPLOADS_BUCKET)
            s3_key = body["s3Key"]

            # Fetch full job record for batch metadata (needed by every mode).
            job = get_job(job_id)
            if not job:
                raise ValueError(f"Job {job_id} not found in DynamoDB")

            file_type = _input_file_type(job, s3_key)
            is_worker = "pageStart" in body
            if file_type in {"csv", "xlsx"}:
                if is_worker:
                    raise ValueError(
                        "Spreadsheet input cannot be processed as a PDF chunk."
                    )
                mode = "tabular"
            elif is_worker:
                mode = "worker"
            elif chunk_pages == 0:
                mode = "serial"  # rollback switch — original serial path (§10)
            else:
                mode = "splitter"

            user_sub = job.get("userSub") or job.get("userId") or "unknown-user"
            batch_id = job.get("batchId", "")
            total_files = int(job.get("totalFilesInBatch", 1))

            # New free-text metadata fields (mandatory for new uploads).
            association = (job.get("association") or "").strip()
            constituency_field = (job.get("constituency") or "").strip()
            council_area = (job.get("councilArea") or "").strip()
            election_label = (job.get("election") or "").strip()
            election_date_field = (job.get("electionDate") or "").strip()

            # Legacy fallback for in-flight jobs created before the schema change.
            constituency_ons = job.get("constituencyOnsCode") or job.get("pconCode", "")
            election_id = job.get("electionId", "")
            constituency_name = (
                constituency_field
                or job.get("constituencyName")
                or constituency_ons
                or "Unknown Constituency"
            )
            election_name = election_label or job.get("electionName") or election_id

            output_prefix = f"outputs/{user_sub}/{batch_id or job_id}"
            combine_payload = {
                "batchId": batch_id,
                "userSub": user_sub,
                "association": association,
                "constituency": constituency_field,
                "councilArea": council_area,
                "election": election_label,
                "electionDate": election_date_field,
                "constituencyOnsCode": constituency_ons,
                "electionId": election_id,
                "totalFilesInBatch": total_files,
            }
            settle_ctx = {
                "mode": mode,
                "batch_id": batch_id,
                "total_files": total_files,
                "combine_payload": combine_payload,
                "output_prefix": output_prefix,
            }

            output_meta = {
                "jobId": job_id,
                "userSub": user_sub,
                "batchId": batch_id,
                "association": association,
                "constituency": constituency_field,
                "councilArea": council_area,
                "election": election_label,
                "electionDate": election_date_field,
                "constituencyOnsCode": constituency_ons,
                "electionId": election_id,
            }

            # ── CSV / XLSX (direct single-chunk job) ────────────────────────
            if mode == "tabular":
                tracker = _ensure_single_chunk_tracker(job_id) or {}
                settled_chunks = tracker.get("settledChunks") or set()
                failed_chunks = tracker.get("failedChunks") or set()

                # A prior delivery may have settled the chunk but failed during
                # the terminal job/batch update. Resume that exact outcome
                # before applying the receive-count cutoff.
                if 0 in settled_chunks:
                    was_failed = 0 in failed_chunks
                    logger.info(
                        "Spreadsheet job %s chunk already settled (%s) — resuming finalisation",
                        job_id, "failed" if was_failed else "succeeded",
                    )
                    try_finalise_job(
                        job_id, 0, batch_id, total_files, combine_payload,
                        output_prefix,
                        chunk_failed=was_failed,
                        failure_reason=(
                            job.get("failureReason")
                            if was_failed
                            else None
                        ),
                    )
                    continue

                # A delayed/redelivered message must not reverse a terminal
                # result. Re-run only the idempotent settlement path so an
                # interrupted batch-count operation can still complete.
                job_status = str(job.get("status") or "").strip().upper()
                if job_status in {"SUCCEEDED", "FAILED"}:
                    logger.info(
                        "Spreadsheet job %s is already %s — retrying settlement only",
                        job_id, job_status,
                    )
                    try_finalise_job(
                        job_id, 0, batch_id, total_files, combine_payload,
                        output_prefix,
                        chunk_failed=(job_status == "FAILED"),
                        failure_reason=job.get("failureReason"),
                    )
                    continue

                receive_count = int(
                    (record.get("attributes") or {}).get(
                        "ApproximateReceiveCount", "1"
                    ) or "1"
                )
                if receive_count >= 3:
                    logger.error(
                        "Spreadsheet job %s at receive count %d — settling as FAILED",
                        job_id, receive_count,
                    )
                    try_finalise_job(
                        job_id, 0, batch_id, total_files, combine_payload,
                        output_prefix, chunk_failed=True,
                        failure_reason="Spreadsheet processing failed after repeated infrastructure errors.",
                    )
                    continue

                suffix = Path(s3_key).suffix or f".{file_type}"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                    tmp_path = tmp.name
                try:
                    input_error_type = (
                        XlsxInputError if file_type == "xlsx" else CsvInputError
                    )
                    input_label = (
                        "Excel workbook" if file_type == "xlsx" else "CSV file"
                    )
                    input_code = file_type.upper()
                    object_info = s3_client.head_object(Bucket=bucket, Key=s3_key)
                    object_size = int(object_info.get("ContentLength") or 0)
                    if object_size <= 0:
                        raise input_error_type(
                            f"{input_code}_EMPTY",
                            f"The {input_label} is empty or contains no data rows.",
                        )
                    if object_size > CSV_MAX_BYTES:
                        raise input_error_type(
                            f"{input_code}_TOO_LARGE",
                            f"The {input_label} exceeds the configured processing size limit.",
                        )
                    s3_client.download_file(bucket, s3_key, tmp_path)
                    parser = (
                        _parse_uploaded_xlsx
                        if file_type == "xlsx"
                        else _parse_uploaded_csv
                    )
                    rows, tabular_meta = parser(
                        tmp_path,
                        constituency_name=constituency_field,
                        election_date=election_date_field,
                    )
                except CsvInputError as input_exc:
                    logger.warning(
                        "Spreadsheet job %s rejected: %s",
                        job_id,
                        input_exc,
                    )
                    try_finalise_job(
                        job_id, 0, batch_id, total_files, combine_payload,
                        output_prefix, chunk_failed=True,
                        failure_reason=str(input_exc),
                    )
                    continue
                finally:
                    _cleanup(tmp_path)

                output_key = f"{output_prefix}/{job_id}.json"
                output_payload = dict(output_meta)
                output_payload.update({
                    "rows": rows,
                    "meta": tabular_meta,
                    "processedAt": datetime.now(timezone.utc).isoformat(),
                })
                s3_client.put_object(
                    Bucket=UPLOADS_BUCKET,
                    Key=output_key,
                    Body=json.dumps(output_payload),
                    ContentType="application/json",
                )
                logger.info(
                    "Spreadsheet job %s wrote %d normalised rows",
                    job_id,
                    len(rows),
                )
                try_finalise_job(
                    job_id, 0, batch_id, total_files, combine_payload,
                    output_prefix, chunk_failed=False,
                )
                continue

            if not OCR_AVAILABLE:
                logger.error(
                    "OCR libraries not available — Tesseract layer may be missing"
                )
                raise RuntimeError("OCR libraries not available")

            # ── Splitter ──────────────────────────────────────────────────────
            if mode == "splitter":
                logger.info("Splitter: job %s (s3://%s/%s)", job_id, bucket, s3_key)
                run_splitter(job_id, bucket, s3_key, chunk_pages)
                continue

            # ── Worker (chunk) ────────────────────────────────────────────────
            if mode == "worker":
                page_start = int(body["pageStart"])
                page_end = int(body["pageEnd"])
                chunk_index = int(body.get("chunkIndex", 0))
                total_chunks = int(body.get("totalChunks", 1))

                receive_count = int(
                    (record.get("attributes") or {}).get("ApproximateReceiveCount", "1") or "1"
                )

                chunk_failed = False
                if receive_count >= 3:
                    # Final delivery before the DLQ — do not attempt the work.
                    # Settle this chunk as failed so the job settles as FAILED and
                    # the batch completes, converting a permanent hang into a
                    # clean, visible failure (§5.7.2).
                    logger.error(
                        "Job %s chunk %d at receive count %d — settling as FAILED (§5.7.2)",
                        job_id, chunk_index, receive_count,
                    )
                    chunk_failed = True
                else:
                    try:
                        logger.info(
                            "Worker: job %s chunk %d/%d pages %d-%d",
                            job_id, chunk_index, total_chunks, page_start, page_end,
                        )
                        suffix = Path(s3_key).suffix or ".pdf"
                        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                            tmp_path = tmp.name
                        s3_client.download_file(bucket, s3_key, tmp_path)
                        logger.info("Downloaded to %s (%d bytes)", tmp_path, os.path.getsize(tmp_path))

                        rows, ocr_meta, page_districts, page_declared_ranges = ocr_pdf(
                            tmp_path,
                            constituency_name,
                            election_name,
                            election_date_override=election_date_field,
                            page_start=page_start,
                            page_end=page_end,
                        )
                        _cleanup(tmp_path)

                        # Zero-pad chunkIndex so lexicographic S3 listing == numeric order.
                        output_key = f"{output_prefix}/{job_id}-{chunk_index:04d}.json"
                        output_payload = dict(output_meta)
                        output_payload.update({
                            "chunkIndex": chunk_index,
                            "totalChunks": total_chunks,
                            "pageDistricts": page_districts,
                            "pageDeclaredRanges": page_declared_ranges,
                            "rows": rows,
                            "meta": ocr_meta,
                            "processedAt": datetime.now(timezone.utc).isoformat(),
                        })
                        s3_client.put_object(
                            Bucket=UPLOADS_BUCKET,
                            Key=output_key,
                            Body=json.dumps(output_payload),
                            ContentType="application/json",
                        )
                        logger.info(
                            "Wrote chunk output s3://%s/%s (%d rows)",
                            UPLOADS_BUCKET, output_key, len(rows),
                        )
                    except Exception as work_exc:
                        # An OCR / download / output failure fails this chunk, and so
                        # the whole file (§5.7.3). It is not retried — a bad page
                        # range or corrupt PDF will not improve on redelivery, and
                        # receive_count already handles genuine timeouts. Settle the
                        # chunk as failed below so the job settles and the batch
                        # completes.
                        logger.exception(
                            "Job %s chunk %d failed during OCR: %s", job_id, chunk_index, work_exc
                        )
                        chunk_failed = True

                # Settle this chunk exactly once for this message. try_finalise_job's
                # tracker updates are idempotent (number sets), so if it raises on a
                # transient DynamoDB error the outer handler can safely return the
                # message to SQS for retry rather than stranding the job.
                try_finalise_job(job_id, chunk_index, batch_id, total_files,
                                 combine_payload, output_prefix, chunk_failed=chunk_failed)
                continue

            # ── Serial (rollback, CHUNK_PAGES=0) ──────────────────────────────
            logger.info("Serial: job %s (s3://%s/%s)", job_id, bucket, s3_key)
            suffix = Path(s3_key).suffix or ".pdf"
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = tmp.name
            s3_client.download_file(bucket, s3_key, tmp_path)
            logger.info("Downloaded to %s (%d bytes)", tmp_path, os.path.getsize(tmp_path))

            rows, ocr_meta, _, _ = ocr_pdf(
                tmp_path,
                constituency_name,
                election_name,
                election_date_override=election_date_field,
            )
            _cleanup(tmp_path)

            output_key = f"{output_prefix}/{job_id}.json"
            output_payload = dict(output_meta)
            output_payload.update({
                "rows": rows,
                "meta": ocr_meta,
                "processedAt": datetime.now(timezone.utc).isoformat(),
            })
            s3_client.put_object(
                Bucket=UPLOADS_BUCKET,
                Key=output_key,
                Body=json.dumps(output_payload),
                ContentType="application/json",
            )
            logger.info("Wrote output to s3://%s/%s (%d rows)", UPLOADS_BUCKET, output_key, len(rows))

            now_iso = datetime.now(timezone.utc).isoformat()
            update_job_succeeded(job_id, output_key, now_iso)
            if batch_id:
                try_trigger_combiner(
                    batch_id, total_files, combine_payload, job_id
                )

        except Exception as exc:
            logger.exception("Failed to process record for job %s: %s", job_id, exc)
            mode_now = settle_ctx["mode"] if settle_ctx else mode
            if isinstance(exc, CombinerClaimPending):
                message_id = record.get("messageId")
                if message_id:
                    failures.append({"itemIdentifier": message_id})
                continue
            if settle_ctx is None:
                # We do not yet know enough about the job to settle it safely.
                # Return the message to SQS for retry instead of ACKing and
                # potentially stranding a worker/chunk or batch.
                message_id = record.get("messageId")
                if message_id:
                    failures.append({"itemIdentifier": message_id})
                continue
            if mode_now in {"worker", "tabular"}:
                # The chunk was not settled (a transient DynamoDB error in
                # try_finalise_job, or a pre-settle error such as get_job). Return
                # the message to SQS so it is retried rather than ACKed and left to
                # strand the job's chunk tracker. Settlement is idempotent
                # (settledChunks is a set), so a retry cannot double-count; after
                # maxReceiveCount the message lands in the DLQ, visible rather than
                # hung. OCR failures never reach here — they settle as a failed
                # chunk above.
                message_id = record.get("messageId")
                if message_id:
                    failures.append({"itemIdentifier": message_id})
                    logger.info("Job %s: returning message to SQS for retry", job_id)
            elif job_id and settle_ctx:
                # Serial or splitter failure: mark the job FAILED and increment the
                # batch counter so the batch still completes (§5.7.1).
                try:
                    update_job_failed(job_id, str(exc), datetime.now(timezone.utc).isoformat())
                except Exception:
                    pass
                if settle_ctx["batch_id"]:
                    try:
                        try_trigger_combiner(
                            settle_ctx["batch_id"], settle_ctx["total_files"],
                            settle_ctx["combine_payload"], job_id,
                        )
                    except Exception:
                        logger.exception("Failed to settle batch for failed job %s", job_id)
            elif job_id:
                # Failure before we loaded the job record; best-effort FAILED mark.
                try:
                    update_job_failed(job_id, str(exc), datetime.now(timezone.utc).isoformat())
                except Exception:
                    pass
            continue

    return {"batchItemFailures": failures}
