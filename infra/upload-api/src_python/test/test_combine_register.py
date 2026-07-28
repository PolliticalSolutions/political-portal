"""Unit tests for the pure functions of the marked-register combiner.

The district resolver (§6.3 / §6.4) is the riskiest new logic in this change and
is pure logic over a page->district map and elector ranges, so it is tested here
directly with synthetic inputs — no Tesseract, no AWS. Test 1b in the spec calls
out exactly these cases: single district, clean two-district split, unreadable
second header, blank page mid-document, and 47/1-style sub-numbered electors.
"""

from copy import deepcopy
from email import policy
from email.parser import BytesParser
from io import BytesIO

import pytest
from openpyxl import load_workbook

import combine_register.handler as c


def _rows(pairs, seed="SEED"):
    """Build rows from (page, elector_number) pairs, all pre-tagged with a seed
    district as the worker would write them."""
    return [
        {"page": page, "elector_number": en, "polling_district": seed,
         "voted": "Y", "election_date": "", "constituency": "", "postal_vote": "N"}
        for page, en in pairs
    ]


def _districts(rows):
    return [r["polling_district"] for r in rows]


# ── _elector_main_number (§6.3) ───────────────────────────────────────────────

class TestElectorMainNumber:
    def test_plain_number(self):
        assert c._elector_main_number("47") == 47

    def test_sub_number_takes_part_before_slash(self):
        # Must NOT become 471 (the _sort_key re.sub approach would do that).
        assert c._elector_main_number("47/1") == 47

    def test_none(self):
        assert c._elector_main_number(None) is None

    def test_empty(self):
        assert c._elector_main_number("") is None

    def test_non_numeric(self):
        assert c._elector_main_number("abc") is None

    def test_whitespace(self):
        assert c._elector_main_number(" 47 ") == 47


# ── Declared-range selection and validation (§5) ─────────────────────────────

class TestResolveDeclaredRanges:
    def test_cover_and_modal_page_header_agreement_is_trusted(self):
        cover = [{"district": "NAA", "start": 1, "end": 926}]
        pages = {
            "3": [{"district": "NAA", "start": 1, "end": 926}],
            "4": [{"district": "NAA", "start": 1, "end": 926}],
            "5": [{"district": "NAA", "start": 1, "end": 926}],
        }
        trusted, issues = c.resolve_declared_ranges(cover, pages)
        assert trusted["NAA"]["start"] == 1
        assert trusted["NAA"]["end"] == 926
        assert trusted["NAA"]["evidence"] == "cover+page_headers"
        assert issues == []

    def test_one_noisy_header_does_not_override_modal_agreement(self):
        cover = [{"district": "NAA", "start": 1, "end": 926}]
        pages = {
            "3": [{"district": "NAA", "start": 1, "end": 926}],
            "4": [{"district": "NAA", "start": 1, "end": 1926}],
            "5": [{"district": "NAA", "start": 1, "end": 926}],
        }
        trusted, issues = c.resolve_declared_ranges(cover, pages)
        assert trusted["NAA"]["end"] == 926
        assert trusted["NAA"]["header_count"] == 2
        assert issues == []

    def test_mis_ocr_cover_disagreement_is_not_trusted(self):
        cover = [{"district": "NAA", "start": 1, "end": 1926}]
        pages = {
            "3": [{"district": "NAA", "start": 1, "end": 926}],
            "4": [{"district": "NAA", "start": 1, "end": 926}],
            "5": [{"district": "NAA", "start": 1, "end": 926}],
        }
        trusted, issues = c.resolve_declared_ranges(cover, pages)
        assert trusted == {}
        assert len(issues) == 1
        assert "disagrees" in issues[0]
        assert "range not trusted" in issues[0]

    def test_repeated_header_only_range_supports_genuine_second_district(self):
        pages = {
            "8": [{"district": "NAB", "start": 50, "end": 400}],
            "9": [{"district": "NAB", "start": 50, "end": 400}],
        }
        trusted, issues = c.resolve_declared_ranges([], pages)
        assert trusted["NAB"]["evidence"] == "repeated_page_headers"
        assert trusted["NAB"]["start"] == 50
        assert issues == []

    def test_single_header_without_cover_is_not_trusted(self):
        pages = {"8": [{"district": "NAB", "start": 50, "end": 400}]}
        trusted, issues = c.resolve_declared_ranges([], pages)
        assert trusted == {}
        assert "appeared only once" in issues[0]

    def test_no_declarations_reports_issue(self):
        trusted, issues = c.resolve_declared_ranges([], {})
        assert trusted == {}
        assert "No declared elector range" in issues[0]


class TestValidateRowsAgainstDeclaredRanges:
    def test_in_range_out_of_range_and_missing_are_reported_without_dropping(self):
        rows = _rows([
            (3, "9"), (3, "10"), (3, "12"), (3, "12/1"),
            (4, "15"), (4, "16"),
        ], seed="NAA")
        before = deepcopy(rows)
        declared = {"NAA": {"district": "NAA", "start": 10, "end": 15}}

        reports, issues = c.validate_rows_against_declared_ranges(rows, declared)

        assert rows == before
        assert issues == []
        assert reports[0]["captured_count"] == 3
        assert reports[0]["missing"] == [11, 13, 14]
        assert reports[0]["out_of_range"] == ["9", "16"]
        assert reports[0]["out_of_range_count"] == 2

    def test_sub_number_uses_part_before_slash(self):
        rows = _rows([(3, "47/1")], seed="NAA")
        declared = {"NAA": {"district": "NAA", "start": 46, "end": 48}}
        reports, _ = c.validate_rows_against_declared_ranges(rows, declared)
        assert reports[0]["captured_count"] == 1
        assert reports[0]["missing"] == [46, 48]
        assert reports[0]["out_of_range"] == []

    def test_register_not_starting_at_one(self):
        rows = _rows([
            (3, "557"), (3, "558"), (4, "560"), (4, "100"), (4, "3752")
        ], seed="TH7")
        declared = {"TH7": {"district": "TH7", "start": 557, "end": 3751}}
        reports, _ = c.validate_rows_against_declared_ranges(rows, declared)
        report = reports[0]
        assert report["declared_count"] == 3195
        assert report["captured_count"] == 3
        assert report["missing"][0] == 559
        assert report["missing"][-1] == 3751
        assert report["out_of_range"] == ["100", "3752"]

    def test_rows_in_district_without_trusted_range_are_flagged(self):
        rows = _rows([(3, "1")], seed="NAB")
        declared = {"NAA": {"district": "NAA", "start": 1, "end": 2}}
        _, issues = c.validate_rows_against_declared_ranges(rows, declared)
        assert issues == [
            "No trusted declared range was available for extracted district(s): NAB."
        ]


# ── resolve_job_districts (§6.3 / §6.4) ───────────────────────────────────────

class TestResolveJobDistricts:
    # District boundaries are accepted only on corroborated printed header codes.
    # The elector-number-reset trigger ("Layer 3b") was removed in Defect A because
    # it split clean single-district registers on OCR artefacts (house numbers
    # misread as low elector numbers) and on out-of-sequence late-registration
    # electors. These tests pin the header-only behaviour and, for the reset
    # scenarios the old code mishandled, assert that a numeric dip no longer splits.

    def test_single_district_no_signal_returns_seed_everywhere(self):
        """Invariant 1/7: no per-page signal → every row gets the seed."""
        rows = _rows([(3, "1"), (3, "2"), (4, "3"), (4, "4")])
        synthetic = c.resolve_job_districts(rows, {}, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_clean_two_district_split_with_headers(self):
        rows = _rows([(3, "10"), (3, "55"), (4, "56"), (4, "60"),
                      (5, "1"), (5, "5"), (6, "6"), (6, "40")])
        page_districts = {"3": "LA", "4": "LA", "5": "LB", "6": "LB"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert _districts(rows) == ["LA", "LA", "LA", "LA", "LB", "LB", "LB", "LB"]
        assert synthetic == set()

    def test_header_corroboration_splits_on_two_consecutive_pages(self):
        """A different printed code on two consecutive pages is the boundary."""
        rows = _rows([(3, "10"), (3, "55"), (4, "56"), (4, "60"), (5, "61"), (5, "70")])
        page_districts = {"3": "LA", "4": "LB", "5": "LB"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        # Boundary lands on the first page of the two-page LB run (page 4).
        assert _districts(rows) == ["LA", "LA", "LB", "LB", "LB", "LB"]
        assert synthetic == set()

    def test_one_off_header_noise_is_suppressed(self):
        """A single stray code (not on two consecutive pages) must not split."""
        rows = _rows([(3, "10"), (3, "20"), (4, "21"), (4, "30"), (5, "31"), (5, "40")])
        page_districts = {"3": "LA", "4": "LX", "5": "LA"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_alternating_header_noise_is_not_corroboration(self):
        """A repeated error cannot corroborate across a different readable code."""
        rows = _rows([
            (3, "10"), (4, "11"), (5, "12"), (6, "13"), (7, "14"),
        ])
        page_districts = {
            "3": "LA", "4": "LA", "5": "LX", "6": "LA", "7": "LX",
        }
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_header_split_midway_through_file(self):
        """A printed code change midway (LA→LB on two consecutive pages) splits;
        elector numbers are irrelevant to the decision."""
        rows = _rows([(3, "55"), (4, "60"), (5, "1"), (6, "40")])
        page_districts = {"3": "LA", "4": "LA", "5": "LB", "6": "LB"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert _districts(rows) == ["LA", "LA", "LB", "LB"]
        assert synthetic == set()

    def test_blank_page_mid_document_is_not_a_boundary(self):
        """A page with no rows is simply skipped; the run continues unbroken."""
        rows = _rows([(3, "10"), (3, "55"), (4, "56"), (4, "60"), (6, "61"), (6, "70")])
        page_districts = {"3": "LA", "4": "LA", "5": "LA", "6": "LA"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_corroborated_cover_header_establishes_short_district_boundary(self):
        """A cover-page header can be the first half of corroboration even when
        that page has no elector rows. This is the Stafford ECG transition:
        page 63 and page 65 print ECG, while only pages 64-65 bear rows."""
        rows = _rows([(64, "1"), (65, "2")])
        synthetic, report = c._resolve_job_districts_with_report(
            rows,
            {"63": "ECG", "64": None, "65": "ECG"},
            "ECF",
        )

        assert _districts(rows) == ["ECG", "ECG"]
        assert synthetic == set()
        assert report["trusted"] is True
        assert report["accepted_districts"] == ["ECG"]
        assert report["unresolved_leading_pages"] == 0

    def test_sub_numbered_electors_do_not_split(self):
        rows = _rows([(3, "47"), (3, "47/1"), (3, "48"), (4, "49"), (4, "50")])
        page_districts = {"3": "LA", "4": "LA"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    # ── Defect A regressions: numeric resets must NOT split a single district ──

    def test_house_number_dip_does_not_split_single_district(self):
        """The reported Defect A regression, reproduced. A run of misread house
        numbers (1..5) drops the page's numbers far below the previous page — the
        exact pattern the old median-reset trigger split into a phantom 'LA-2'.
        With a consistent printed header, and now under header-only logic, it must
        stay a single district."""
        rows = _rows([
            (3, "410"), (3, "411"), (3, "412"), (3, "413"), (3, "414"), (3, "415"),
            (4, "1"), (4, "2"), (4, "3"), (4, "4"), (4, "5"),
            (4, "416"), (4, "417"), (4, "418"),
        ])
        page_districts = {"3": "LA", "4": "LA"}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_full_numeric_reset_without_header_no_longer_splits(self):
        """A full numeric reset (400s → low tens) with no readable header code no
        longer creates a synthetic district — the old code produced 'LA-2' here.
        Without a printed code we do not invent a boundary; the run stays 'LA'."""
        rows = _rows([
            (3, "450"), (3, "460"), (3, "470"), (3, "480"), (3, "490"),
            (4, "1"), (4, "5"), (4, "10"), (4, "15"), (4, "20"),
        ])
        page_districts = {"3": "LA", "4": None}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_unreadable_second_header_no_longer_splits(self):
        """Deliberately accepted trade-off: a genuine second district whose header
        cannot be OCR'd is no longer detected by a numeric reset, so its rows
        inherit the running district here. In the real pipeline this collision
        surfaces downstream as a high dedupe rate and the customer result is
        withheld, rather than creating a synthetic label. This test pins the
        intentional behaviour."""
        rows = _rows([(3, "10"), (3, "55"), (4, "56"), (4, "60"),
                      (5, "1"), (5, "5"), (6, "6"), (6, "40")])
        page_districts = {"3": "LA", "4": "LA", "5": None, "6": None}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_repeated_numeric_resets_without_headers_do_not_split(self):
        """Several numeric resets across a headerless run must not manufacture a
        cascade of synthetic districts (the old code produced LA-2, LA-3, ...)."""
        rows = _rows([(3, "10"), (3, "55"), (4, "56"), (4, "60"),
                      (5, "1"), (5, "5"), (6, "6"), (6, "60"),
                      (7, "1"), (7, "5"), (8, "6"), (8, "40")])
        page_districts = {str(p): None for p in range(3, 9)}
        synthetic = c.resolve_job_districts(rows, page_districts, "LA")
        assert set(_districts(rows)) == {"LA"}
        assert synthetic == set()

    def test_empty_rows(self):
        assert c.resolve_job_districts([], {"3": "LA"}, "LA") == set()

    def test_rows_without_page_are_left_untouched(self):
        rows = [{"elector_number": "1", "polling_district": "LA"}]
        synthetic = c.resolve_job_districts(rows, {}, "SEED")
        # No page → not grouped → not reassigned.
        assert rows[0]["polling_district"] == "LA"
        assert synthetic == set()

    def test_corroborated_headers_produce_trusted_resolution_report(self):
        rows = _rows([
            (3, "1"), (3, "2"),
            (4, "3"), (4, "4"),
            (5, "5"), (5, "6"),
        ])
        synthetic, report = c._resolve_job_districts_with_report(
            rows,
            {"3": "ECA", "4": "ECA", "5": None},
            "DIVISION",
        )

        assert synthetic == set()
        assert report["trusted"] is True
        assert report["accepted_districts"] == ["ECA"]
        assert report["header_coverage_pct"] == 66.7
        assert report["unresolved_leading_pages"] == 0
        assert set(_districts(rows)) == {"ECA"}

    def test_corroboration_tolerates_one_missed_header_page(self):
        rows = _rows([(3, "1"), (4, "2"), (5, "3")])
        _synthetic, report = c._resolve_job_districts_with_report(
            rows,
            {"3": "ECA", "4": None, "5": "ECA"},
            "DIVISION",
        )

        assert report["trusted"] is True
        assert report["accepted_districts"] == ["ECA"]
        assert set(_districts(rows)) == {"ECA"}

    def test_single_uncorroborated_header_is_not_trusted(self):
        rows = _rows([(3, "1"), (4, "2"), (5, "3")])
        _synthetic, report = c._resolve_job_districts_with_report(
            rows,
            {"3": "ECA", "4": None, "5": None},
            "DIVISION",
        )

        assert report["trusted"] is False
        assert report["accepted_districts"] == []
        assert report["rows_with_untrusted_district"] == 3
        assert any("within the next two pages" in issue for issue in report["issues"])

    def test_low_header_coverage_is_not_trusted_even_with_one_good_run(self):
        rows = _rows([(page, str(page)) for page in range(3, 18)])
        _synthetic, report = c._resolve_job_districts_with_report(
            rows,
            {"3": "ECA", "4": "ECA"},
            "ECA",
        )

        assert report["header_coverage_pct"] == 13.3
        assert report["trusted"] is False
        assert any("minimum 20%" in issue for issue in report["issues"])


# ── _dedupe_rows (§6.1 — the core bug this fixes) ─────────────────────────────

class TestDedupeRows:
    def test_cross_district_overlap_preserved(self):
        rows = [
            {"polling_district": "LA", "elector_number": "47"},
            {"polling_district": "LB", "elector_number": "47"},
        ]
        out = c._dedupe_rows(rows)
        assert len(out) == 2

    def test_within_district_duplicate_removed(self):
        rows = [
            {"polling_district": "LA", "elector_number": "47"},
            {"polling_district": "LA", "elector_number": "47"},
        ]
        out = c._dedupe_rows(rows)
        assert len(out) == 1

    def test_empty_elector_number_skipped(self):
        rows = [{"polling_district": "LA", "elector_number": ""}]
        assert c._dedupe_rows(rows) == []

    def test_same_source_conflict_preserves_first_pdf_row(self):
        rows = [
            {
                "polling_district": "LA", "elector_number": "1",
                "voted": "N", "postal_vote": "N", "_source_type": "pdf",
            },
            {
                "polling_district": "LA", "elector_number": "1",
                "voted": "Y", "postal_vote": "Y", "_source_type": "pdf",
            },
        ]
        out = c._dedupe_rows(rows)
        assert out[0]["voted"] == "N"
        assert out[0]["postal_vote"] == "N"

    def test_pdf_vote_and_csv_postal_status_are_merged_independently(self):
        rows = [
            {
                "polling_district": "LA", "elector_number": "47/1",
                "voted": "Y", "postal_vote": "N", "_source_type": "pdf",
            },
            {
                "polling_district": "LA", "elector_number": "47/1",
                "voted": "N", "postal_vote": "Y", "_source_type": "csv",
            },
        ]
        assert c._dedupe_rows(rows) == [{
            "polling_district": "LA",
            "elector_number": "47/1",
            "voted": "Y",
            "postal_vote": "Y",
            "_source_type": "pdf",
        }]

    def test_flag_merge_is_order_independent(self):
        pdf = {
            "polling_district": "LA", "elector_number": "47/1",
            "voted": "Y", "postal_vote": "N", "_source_type": "pdf",
        }
        absent_voter = {
            "polling_district": "LA", "elector_number": "47/1",
            "voted": "N", "postal_vote": "Y", "_source_type": "csv",
        }
        pdf_first = c._dedupe_rows([pdf, absent_voter])[0]
        csv_first = c._dedupe_rows([absent_voter, pdf])[0]
        assert (pdf_first["voted"], pdf_first["postal_vote"]) == ("Y", "Y")
        assert (csv_first["voted"], csv_first["postal_vote"]) == ("Y", "Y")

    def test_cross_source_key_normalises_district_case_and_outer_whitespace(self):
        pdf = {
            "election_date": "01/05/2026",
            "constituency": "Test",
            "polling_district": " pd1 ",
            "elector_number": " 47/1 ",
            "voted": "Y",
            "postal_vote": "N",
            "_source_type": "pdf",
        }
        absent_voter = {
            "election_date": "01/05/2026",
            "constituency": "Test",
            "polling_district": "PD1",
            "elector_number": "47/1",
            "voted": "N",
            "postal_vote": "Y",
            "_source_type": "csv",
        }
        pdf_first = c._dedupe_rows([pdf, absent_voter])
        csv_first = c._dedupe_rows([absent_voter, pdf])
        assert len(pdf_first) == 1
        assert c.build_csv(pdf_first) == c.build_csv(csv_first)
        assert pdf_first[0]["polling_district"] == "PD1"
        assert pdf_first[0]["elector_number"] == "47/1"
        assert (pdf_first[0]["voted"], pdf_first[0]["postal_vote"]) == ("Y", "Y")

    def test_subnumbers_remain_distinct_keys(self):
        rows = [
            {"polling_district": "LA", "elector_number": "47"},
            {"polling_district": "LA", "elector_number": "47/1"},
            {"polling_district": "LA", "elector_number": "47/2"},
        ]
        assert len(c._dedupe_rows(rows)) == 3

    def test_dedupe_does_not_mutate_input_rows(self):
        rows = [
            {
                "polling_district": "LA", "elector_number": "1",
                "voted": "N", "postal_vote": "N",
            },
            {
                "polling_district": "LA", "elector_number": "1",
                "voted": "Y", "postal_vote": "Y",
            },
        ]
        before = deepcopy(rows)
        c._dedupe_rows(rows)
        assert rows == before


class TestDedupeSourceCounts:
    def test_cross_source_match_is_not_a_within_source_duplicate(self):
        rows = [
            {
                "polling_district": "LA", "elector_number": "1",
                "_source_type": "pdf",
            },
            {
                "polling_district": "LA", "elector_number": "1",
                "_source_type": "csv",
            },
        ]
        assert c._dedupe_source_counts(rows) == {
            "within_source": 0,
            "cross_source": 1,
            "total": 1,
        }

    def test_same_source_repeat_remains_a_duplicate(self):
        rows = [
            {
                "polling_district": "LA", "elector_number": "1",
                "_source_type": "pdf",
            },
            {
                "polling_district": "LA", "elector_number": "1",
                "_source_type": "pdf",
            },
            {
                "polling_district": "LA", "elector_number": "1",
                "_source_type": "csv",
            },
        ]
        assert c._dedupe_source_counts(rows) == {
            "within_source": 1,
            "cross_source": 1,
            "total": 2,
        }

    def test_source_counts_use_the_same_canonical_district_key(self):
        rows = [
            {
                "polling_district": " pd1 ",
                "elector_number": "1",
                "_source_type": "pdf",
            },
            {
                "polling_district": "PD1",
                "elector_number": "1",
                "_source_type": "csv",
            },
        ]
        assert c._dedupe_source_counts(rows) == {
            "within_source": 0,
            "cross_source": 1,
            "total": 1,
        }


# ── _sort_key ─────────────────────────────────────────────────────────────────

class TestSortKey:
    def test_sorts_by_district_then_numeric_elector(self):
        rows = [
            {"polling_district": "LB", "elector_number": "2"},
            {"polling_district": "LA", "elector_number": "10"},
            {"polling_district": "LA", "elector_number": "2"},
        ]
        rows.sort(key=c._sort_key)
        assert [(r["polling_district"], r["elector_number"]) for r in rows] == [
            ("LA", "2"), ("LA", "10"), ("LB", "2"),
        ]


# ── build_filename ────────────────────────────────────────────────────────────

class TestBuildFilename:
    def test_all_parts_joined(self):
        # '/' is a forbidden filename char and is sanitised to a space.
        name = c.build_filename("Assoc", "Const", "Council", "GE 2026", "01/05/2026", "batch-1")
        assert name == "Assoc - Const - Council - GE 2026 - 01 05 2026 - Marked Register.xlsx"

    def test_missing_part_uses_fallback(self):
        name = c.build_filename("Assoc", "", "Council", "GE 2026", "01/05/2026", "batch-1")
        assert name == "batch-1 - Marked Register.xlsx"

    def test_forbidden_characters_sanitised(self):
        name = c.build_filename("A/B", "C:D", "E", "F", "G", "batch")
        assert "/" not in name.replace(" - ", "")
        assert ":" not in name


# ── District reporting + warnings (§6.5) ──────────────────────────────────────

class TestDistrictReporting:
    def test_count_districts(self):
        rows = [
            {"polling_district": "LA"}, {"polling_district": "LA"},
            {"polling_district": "LB"},
        ]
        assert c._count_districts(rows) == {"LA": 2, "LB": 1}

    def test_format_districts_orders_by_count_desc(self):
        text = c._format_districts({"LA": 1204, "LB": 987, "LA-2": 640})
        assert text == "Polling districts: 3 (LA: 1,204, LB: 987, LA-2: 640)"

    def test_format_districts_empty(self):
        assert c._format_districts({}) == "Polling districts: 0"


class TestWarningsTriggered:
    def test_high_dedupe_pct_triggers(self):
        assert c._warnings_triggered(dedupe_pct=30.0, synthetic_labels=set(), warn_pct=2.0) is True

    def test_synthetic_label_triggers(self):
        assert c._warnings_triggered(dedupe_pct=0.0, synthetic_labels={"LA-2"}, warn_pct=2.0) is True

    def test_clean_run_no_warning(self):
        assert c._warnings_triggered(dedupe_pct=0.5, synthetic_labels=set(), warn_pct=2.0) is False

    def test_pct_exactly_at_threshold_no_warning(self):
        assert c._warnings_triggered(dedupe_pct=2.0, synthetic_labels=set(), warn_pct=2.0) is False

    def test_missing_declared_electors_trigger_warning(self):
        reports = [{"missing_count": 1, "out_of_range_count": 0, "unparseable_count": 0}]
        assert c._warnings_triggered(
            dedupe_pct=0.0, synthetic_labels=set(), warn_pct=2.0,
            range_reports=reports,
        ) is True

    def test_out_of_range_electors_trigger_warning(self):
        reports = [{"missing_count": 0, "out_of_range_count": 1, "unparseable_count": 0}]
        assert c._warnings_triggered(
            dedupe_pct=0.0, synthetic_labels=set(), warn_pct=2.0,
            range_reports=reports,
        ) is True

    def test_untrusted_range_issue_triggers_warning(self):
        assert c._warnings_triggered(
            dedupe_pct=0.0, synthetic_labels=set(), warn_pct=2.0,
            range_issues=["NAA: range not trusted"],
        ) is True

    def test_complete_declared_range_is_clean(self):
        reports = [{"missing_count": 0, "out_of_range_count": 0, "unparseable_count": 0}]
        assert c._warnings_triggered(
            dedupe_pct=0.0, synthetic_labels=set(), warn_pct=2.0,
            range_reports=reports,
        ) is False


class TestQualityBlockers:
    def test_high_within_source_dedupe_blocks_release(self):
        blockers = c._quality_blockers(
            dedupe_pct=84.3,
            warn_pct=2.0,
            district_counts={"ECA": 100},
            district_resolution_reports=[{"trusted": True}],
        )
        assert blockers == [
            "Deduplication would remove 84.3% of within-source rows "
            "(maximum 2%)."
        ]

    def test_untrusted_resolution_blocks_release(self):
        blockers = c._quality_blockers(
            dedupe_pct=0.0,
            warn_pct=2.0,
            district_counts={"ECA": 100},
            district_resolution_reports=[{
                "trusted": False,
                "source": "register.pdf",
                "issues": ["No district headers were corroborated."],
            }],
        )
        assert blockers == [
            "register.pdf: No district headers were corroborated."
        ]

    def test_clean_resolution_has_no_blockers(self):
        assert c._quality_blockers(
            dedupe_pct=0.5,
            warn_pct=2.0,
            district_counts={"ECA": 100},
            district_resolution_reports=[{"trusted": True}],
        ) == []


class TestRangeReportFormatting:
    def test_numbering_span_observed_and_not_observed_lists(self):
        text = c._format_range_report({
            "source": "register.pdf",
            "district": "NAA",
            "start": 1,
            "end": 926,
            "captured_count": 812,
            "declared_count": 926,
            "captured_pct": 812 / 926 * 100,
            "missing_count": 3,
            "missing": [4, 19, 926],
            "out_of_range": [927],
        })
        assert text == (
            "register.pdf: Declared numbering span: NAA 1-926\n"
            "    Unique base numbers observed within span: 812\n"
            "    Numbers not observed within span (3): [4, 19, 926]\n"
            "    Observed outside the declared span: [927]"
        )

    def test_completion_email_explains_span_is_only_a_review_checklist(self, monkeypatch):
        sent = {}

        class FakeSes:
            def send_raw_email(self, **kwargs):
                sent.update(kwargs)

        monkeypatch.setattr(c, "ses", FakeSes())
        c.send_completion_email(
            filename="synthetic.csv",
            csv_bytes=b"header\n",
            succeeded_count=1,
            failed_count=0,
            failed_filenames=[],
            row_count=812,
            range_reports=[{
                "source": "synthetic.pdf",
                "district": "NAA",
                "start": 1,
                "end": 926,
                "captured_count": 812,
                "declared_count": 926,
                "captured_pct": 812 / 926 * 100,
                "missing_count": 3,
                "missing": [4, 19, 926],
                "out_of_range_count": 1,
                "out_of_range": [927],
                "unparseable_count": 0,
            }],
        )

        message = BytesParser(policy=policy.default).parsebytes(
            sent["RawMessage"]["Data"]
        )
        body = message.get_body(preferencelist=("plain",)).get_content()
        assert "Declared numbering span: NAA 1-926" in body
        assert "Unique base numbers observed within span: 812" in body
        assert "Numbers not observed within span (3): [4, 19, 926]" in body
        assert "Observed outside the declared span: [927]" in body
        assert (
            "Numbering spans may contain legitimate gaps. This section is a "
            "review checklist, not an electorate count, extraction-accuracy "
            "score, or turnout calculation."
        ) in body
        assert "87.7%" not in body
        assert "captured 812 of 926" not in body


class TestCompletionDelivery:
    def test_quality_review_email_has_no_attachment_or_download(self):
        raw, mode = c.prepare_quality_review_email(
            filename="result.csv",
            succeeded_count=7,
            failed_count=0,
            candidate_row_count=65_873,
            quality_blockers=["District resolution failed."],
        )
        message = BytesParser(policy=policy.default).parsebytes(raw)
        body = message.get_body(preferencelist=("plain",)).get_content()

        assert mode == "NOTICE_ONLY"
        assert list(message.iter_attachments()) == []
        assert "No output file was released" in body
        assert "District resolution failed." in body

    def test_small_result_is_attached(self, monkeypatch):
        sent = {}

        class FakeSes:
            def send_raw_email(self, **kwargs):
                sent.update(kwargs)

        monkeypatch.setattr(c, "ses", FakeSes())
        monkeypatch.setattr(c, "SES_MAX_RAW_EMAIL_BYTES", 1_000_000)

        mode = c.send_completion_email(
            filename="synthetic.xlsx",
            csv_bytes=c.build_xlsx([]),
            succeeded_count=1,
            failed_count=0,
            failed_filenames=[],
            row_count=1,
        )

        message = BytesParser(policy=policy.default).parsebytes(
            sent["RawMessage"]["Data"]
        )
        assert mode == "ATTACHMENT"
        assert message.get_body(preferencelist=("plain",)) is not None
        attachments = list(message.iter_attachments())
        assert len(attachments) == 1
        assert attachments[0].get_content_type() == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_oversized_result_uses_authenticated_portal_link(self, monkeypatch):
        sent = {}

        class FakeSes:
            def send_raw_email(self, **kwargs):
                sent.update(kwargs)

        monkeypatch.setattr(c, "ses", FakeSes())
        monkeypatch.setattr(c, "SES_MAX_RAW_EMAIL_BYTES", 1)
        monkeypatch.setattr(
            c, "PLATFORM_BASE_URL", "https://www.politicalsolutions.uk"
        )
        generated_for = []
        monkeypatch.setattr(
            c,
            "generate_download_url",
            lambda key: (
                generated_for.append(key)
                or "https://downloads.example.test/result.csv?temporary=signature"
            ),
        )

        mode = c.send_completion_email(
            filename="large.csv",
            csv_bytes=b"header\n" + (b"value\n" * 100),
            succeeded_count=1,
            failed_count=0,
            failed_filenames=[],
            row_count=100,
            csv_key="outputs/user-1/batch-1/large.csv",
        )

        message = BytesParser(policy=policy.default).parsebytes(
            sent["RawMessage"]["Data"]
        )
        body = message.get_body(preferencelist=("plain",)).get_content()
        assert mode == "DOWNLOAD_LINK"
        assert list(message.iter_attachments()) == []
        assert "too large to send safely as an email attachment" in body
        assert "https://downloads.example.test/result.csv?temporary=signature" in body
        assert "https://www.politicalsolutions.uk/portal/uploads" in body
        assert generated_for == ["outputs/user-1/batch-1/large.csv"]

    def test_email_failure_code_is_sanitised(self):
        class SesError(RuntimeError):
            response = {"Error": {"Code": "InvalidParameterValue"}}

        assert (
            c._completion_email_failure_code(SesError())
            == "SES_INVALIDPARAMETERVALUE"
        )
        assert c._completion_email_failure_code(RuntimeError()) == "EMAIL_SEND_FAILED"

    def test_handler_persists_failure_then_reraises_for_lambda_alarm(
        self, monkeypatch
    ):
        jobs = [{
            "jobId": "job-1",
            "status": "FAILED",
            "filename": "unsupported.xlsx",
        }]
        updates = []

        monkeypatch.setattr(c, "get_all_batch_jobs", lambda _batch_id: jobs)
        monkeypatch.setattr(
            c,
            "upload_csv",
            lambda _user_sub, _batch_id, _filename, _content: (
                "outputs/user-1/batch-1/result.csv"
            ),
        )
        monkeypatch.setattr(
            c,
            "update_batch_jobs",
            lambda updated_jobs, **kwargs: updates.append(
                (updated_jobs, kwargs)
            ),
        )
        monkeypatch.setattr(
            c,
            "send_prepared_completion_email",
            lambda _message: (_ for _ in ()).throw(RuntimeError("SES unavailable")),
        )

        with pytest.raises(RuntimeError, match="SES unavailable"):
            c.handler(
                {
                    "batchId": "batch-1",
                    "userSub": "user-1",
                    "association": "Test Association",
                    "constituency": "Test Constituency",
                    "councilArea": "Test Council",
                    "election": "Test Election",
                    "electionDate": "01 May 2026",
                },
                None,
            )

        assert [update[1]["email_status"] for update in updates] == [
            "PENDING",
            "FAILED",
        ]
        failed = updates[-1][1]
        assert failed["batch_status"] == "COMPLETE_WITH_FAILURES"
        assert failed["email_failure_code"] == "EMAIL_SEND_FAILED"
        assert failed["output_key"] == "outputs/user-1/batch-1/result.csv"

    def test_handler_skips_duplicate_email_after_acceptance_marker(
        self, monkeypatch
    ):
        jobs = [{
            "jobId": "job-1",
            "status": "FAILED",
            "filename": "unsupported.xlsx",
            "completionEmailStatus": "SENT",
            "completionEmailMode": "DOWNLOAD_LINK",
            "completionEmailUpdatedAt": "2026-07-24T18:00:00+00:00",
            "batchOutputKey": "outputs/user-1/batch-1/result.csv",
        }]
        updates = []
        send_calls = []

        monkeypatch.setattr(c, "get_all_batch_jobs", lambda _batch_id: jobs)
        monkeypatch.setattr(
            c,
            "upload_csv",
            lambda _user_sub, _batch_id, _filename, _content: (
                "outputs/user-1/batch-1/result.csv"
            ),
        )
        monkeypatch.setattr(
            c,
            "update_batch_jobs",
            lambda updated_jobs, **kwargs: updates.append(
                (updated_jobs, kwargs)
            ),
        )
        monkeypatch.setattr(
            c,
            "send_prepared_completion_email",
            lambda message: send_calls.append(message),
        )

        result = c.handler(
            {
                "batchId": "batch-1",
                "userSub": "user-1",
                "association": "Test Association",
                "constituency": "Test Constituency",
                "councilArea": "Test Council",
                "election": "Test Election",
                "electionDate": "01 May 2026",
            },
            None,
        )

        assert send_calls == []
        assert len(updates) == 1
        assert updates[0][1]["email_status"] == "SENT"
        assert updates[0][1]["email_mode"] == "DOWNLOAD_LINK"
        assert result["completionEmailMode"] == "DOWNLOAD_LINK"

    def test_handler_withholds_high_dedupe_pdf_without_uploading_csv(
        self, monkeypatch
    ):
        jobs = [{
            "jobId": "job-1",
            "status": "SUCCEEDED",
            "filename": "register.pdf",
        }]
        payloads = [{
            "meta": {
                "source_type": "pdf",
                "polling_district": "ECA",
            },
            "rows": [
                {
                    "page": 3, "polling_district": "ECA",
                    "elector_number": "1", "voted": "N",
                    "postal_vote": "N",
                },
                {
                    "page": 4, "polling_district": "ECA",
                    "elector_number": "2", "voted": "N",
                    "postal_vote": "N",
                },
                {
                    "page": 5, "polling_district": "ECA",
                    "elector_number": "1", "voted": "Y",
                    "postal_vote": "N",
                },
            ],
            "pageDistricts": {"3": "ECA", "4": "ECA", "5": "ECA"},
        }]
        updates = []
        marker_updates = []
        sent = []

        monkeypatch.setattr(c, "get_all_batch_jobs", lambda _batch_id: jobs)
        monkeypatch.setattr(
            c,
            "read_job_outputs",
            lambda _user_sub, _batch_id, _job_id: payloads,
        )
        monkeypatch.setattr(
            c,
            "upload_csv",
            lambda *_args, **_kwargs: pytest.fail(
                "A blocked batch must not upload a result."
            ),
        )
        monkeypatch.setattr(
            c,
            "update_batch_jobs",
            lambda updated_jobs, **kwargs: updates.append(
                (updated_jobs, kwargs)
            ),
        )
        monkeypatch.setattr(
            c,
            "update_job_batch_completion",
            lambda job_id, **kwargs: marker_updates.append((job_id, kwargs)),
        )
        monkeypatch.setattr(
            c,
            "send_prepared_completion_email",
            lambda message: sent.append(message),
        )

        result = c.handler(
            {
                "batchId": "batch-1",
                "userSub": "user-1",
                "association": "Test Association",
                "constituency": "Test Constituency",
                "councilArea": "Test Council",
                "election": "Test Election",
                "electionDate": "01 May 2026",
            },
            None,
        )

        assert result["batchStatus"] == "QUALITY_REVIEW_REQUIRED"
        assert result["completionEmailMode"] == "NOTICE_ONLY"
        assert result["csvKey"] == ""
        assert result["qualityBlockerCount"] == 1
        assert len(sent) == 1
        assert [update[1]["email_status"] for update in updates] == [
            "PENDING",
            "SENT",
        ]
        assert marker_updates[0][1]["output_key"] == ""


# ── build_csv column mapping (page field must never reach the CSV) ────────────

class TestBuildCsv:
    def test_page_field_not_in_output(self):
        rows = [{
            "election_date": "01/05/2026", "constituency": "C", "polling_district": "LA",
            "elector_number": "47", "voted": "Y", "postal_vote": "N", "page": 3,
        }]
        out = c.build_csv(rows)
        header = out.splitlines()[0]
        assert header == "Election Date,Constituency,Polling District,Elector Number,Voted,Postal Vote"
        assert "page" not in out.lower().splitlines()[0]

    def test_header_and_row(self):
        rows = [{
            "election_date": "01/05/2026", "constituency": "C", "polling_district": "LA",
            "elector_number": "47", "voted": "Y", "postal_vote": "N",
        }]
        out = c.build_csv(rows)
        lines = out.strip().splitlines()
        assert lines[1] == "01/05/2026,C,LA,47,Y,N"


class TestBuildXlsx:
    def test_roll_number_is_literal_text_not_a_date(self):
        rows = [{
            "election_date": "01/05/2026",
            "constituency": "C",
            "polling_district": "LA",
            "elector_number": "12/3",
            "voted": "Y",
            "postal_vote": "N",
        }]

        workbook = load_workbook(BytesIO(c.build_xlsx(rows)), data_only=False)
        worksheet = workbook["Marked Register"]

        assert worksheet["D2"].value == "12/3"
        assert worksheet["D2"].data_type == "s"
        assert worksheet["D2"].number_format == "@"

    def test_formula_shaped_value_is_literal_text(self):
        rows = [{
            "election_date": "",
            "constituency": "",
            "polling_district": "LA",
            "elector_number": "=1+1",
            "voted": "N",
            "postal_vote": "N",
        }]

        workbook = load_workbook(BytesIO(c.build_xlsx(rows)), data_only=False)
        assert workbook["Marked Register"]["D2"].value == "=1+1"
        assert workbook["Marked Register"]["D2"].data_type == "s"
