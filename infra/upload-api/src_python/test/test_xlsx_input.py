"""Synthetic, privacy-safe tests for XLSX marked-register ingestion."""

import io
import json
import zipfile
from datetime import date
from pathlib import Path

import pytest

import process_register.handler as h

# Import after the handler has selected openpyxl's hardened XML mode.
import openpyxl
from openpyxl import Workbook


XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
PV_MARKED_HEADERS = [
    "RecNo",
    "ElectionDescription",
    "ElectionDate",
    "TimeFrom",
    "TimeTo",
    "PollingName",
    "PollingAddress1",
    "PollingAddress2",
    "PollingAddress3",
    "PollingAddress4",
    "PollingAddress5",
    "ElectorNo",
    "ElectorName",
    "ElectorAddress1",
    "ElectorAddress2",
    "ElectorAddress3",
    "ElectorAddress4",
    "ElectorAddress5",
    "ElectorAddress6",
    "PostalName",
    "PostalAddress1",
    "PostalAddress2",
    "PostalAddress3",
    "PostalAddress4",
    "PostalAddress5",
    "PostalAddress6",
    "PostalAddressPostcode",
    "AreaName1",
    "PVSStatus",
    "DecReceiptDate",
]


def _save_workbook(tmp_path, workbook, name="input.xlsx"):
    path = tmp_path / name
    workbook.save(path)
    workbook.close()
    return path


def _workbook_bytes(workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _populate_flat_sheet(
    worksheet,
    *,
    district="PD1",
    elector_number="47/1",
    marker=True,
    private_value="PRIVATE_DETAIL_REMOVED",
):
    worksheet.append([
        "CompanyName",
        "DistrictRef",
        "ElectorShortNumber",
        "MarkerPostal",
        "ElectorFullName",
    ])
    worksheet.append([
        "Synthetic Council",
        district,
        elector_number,
        marker,
        private_value,
    ])


def _flat_workbook(**kwargs):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Register"
    _populate_flat_sheet(worksheet, **kwargs)
    return workbook


def _populate_report_sheet(
    worksheet,
    *,
    private_detail="PRIVATE_DETAIL_REMOVED",
    footer_date=date(2026, 3, 9),
):
    rows = [
        ["Absent Voter Postal List Marked", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "Date", "", "", ""],
        ["Synthetic election metadata", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        [
            "Reg. No",
            "",
            "Electors Name and Register Address\nPostal Address (if different)",
            "",
            "Ward",
            "",
            "",
            "",
            "",
        ],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["PD1-47/1", "", "", "", "", "Synthetic Ward", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        [private_detail, "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["Total Number of Postal Voters 1", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        [footer_date, "", "", "", "", "", "", "", "Page 1-of-1"],
    ]
    for row in rows:
        worksheet.append(row)


def _report_workbook(**kwargs):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Postal report"
    _populate_report_sheet(worksheet, **kwargs)
    return workbook


def _pv_marked_row(
    *,
    record_number="1",
    elector_number="1A-47/1",
    status="0",
    receipt_date="",
    private_value="PRIVATE_DETAIL_REMOVED",
):
    row = [""] * len(PV_MARKED_HEADERS)
    row[0] = record_number
    row[1] = "Synthetic election"
    row[2] = "Thursday 9 July 2026"
    row[3] = "07:00"
    row[4] = "22:00"
    row[5] = private_value
    row[6] = private_value
    row[11] = elector_number
    row[12] = private_value
    row[13] = private_value
    row[19] = private_value
    row[20] = private_value
    row[27] = "Synthetic area"
    row[28] = status
    row[29] = receipt_date
    return row


def _pv_marked_workbook(*rows, headers=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PV Marked Register"
    worksheet.append(headers or PV_MARKED_HEADERS)
    for row in rows or (_pv_marked_row(),):
        worksheet.append(row)
    return workbook


class TestXlsxParsing:
    def test_flat_profile_matches_csv_shape_and_discards_private_cells(
        self, tmp_path
    ):
        path = _save_workbook(tmp_path, _flat_workbook())

        rows, meta = h._parse_uploaded_xlsx(
            path,
            "Test Constituency",
            "01/05/2026",
        )

        assert rows == [{
            "election_date": "01/05/2026",
            "constituency": "Test Constituency",
            "polling_district": "PD1",
            "elector_number": "47/1",
            "voted": "N",
            "postal_vote": "Y",
        }]
        assert "PRIVATE_DETAIL_REMOVED" not in json.dumps(rows)
        assert meta["source_type"] == "csv"
        assert meta["source_format"] == "xlsx"
        assert meta["csv_schema"] == "absent_voters_v1"

    def test_boolean_and_integral_numeric_cells_are_canonicalised(
        self, tmp_path
    ):
        path = _save_workbook(
            tmp_path,
            _flat_workbook(elector_number=47.0, marker=False),
        )

        rows, _ = h._parse_uploaded_xlsx(path, "Test", "Date")

        assert rows[0]["elector_number"] == "47"
        assert rows[0]["postal_vote"] == "N"

    def test_non_integral_elector_number_is_rejected_privacy_safely(
        self, tmp_path
    ):
        path = _save_workbook(
            tmp_path,
            _flat_workbook(elector_number=47.5),
        )

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_ROW_INVALID"
        assert "47.5" not in str(exc.value)
        assert "rows with invalid ElectorShortNumber: 1" in str(exc.value)

    def test_marked_report_supports_a_typed_footer_date(self, tmp_path):
        path = _save_workbook(tmp_path, _report_workbook())

        rows, meta = h._parse_uploaded_xlsx(
            path,
            "Test Constituency",
            "01/05/2026",
        )

        assert rows == [{
            "election_date": "01/05/2026",
            "constituency": "Test Constituency",
            "polling_district": "PD1",
            "elector_number": "47/1",
            "voted": "N",
            "postal_vote": "Y",
        }]
        rendered = json.dumps({"rows": rows, "meta": meta})
        assert "PRIVATE_DETAIL_REMOVED" not in rendered
        assert "Synthetic Ward" not in rendered
        assert meta["source_type"] == "csv"
        assert meta["source_format"] == "xlsx"
        assert meta["csv_schema"] == "marked_postal_report_v1"

    def test_pv_marked_register_maps_returned_and_unreturned_rows_privately(
        self, tmp_path
    ):
        path = _save_workbook(
            tmp_path,
            _pv_marked_workbook(
                _pv_marked_row(),
                _pv_marked_row(
                    record_number="2",
                    elector_number="1A-48",
                    status="1",
                    receipt_date="09/07/2026 21:14:05",
                ),
            ),
        )

        rows, meta = h._parse_uploaded_xlsx(
            path,
            "Test Constituency",
            "09/07/2026",
        )

        assert rows == [
            {
                "election_date": "09/07/2026",
                "constituency": "Test Constituency",
                "polling_district": "1A",
                "elector_number": "47/1",
                "voted": "N",
                "postal_vote": "Y",
            },
            {
                "election_date": "09/07/2026",
                "constituency": "Test Constituency",
                "polling_district": "1A",
                "elector_number": "48",
                "voted": "Y",
                "postal_vote": "Y",
            },
        ]
        rendered = json.dumps({"rows": rows, "meta": meta})
        assert "PRIVATE_DETAIL_REMOVED" not in rendered
        assert meta["source_type"] == "csv"
        assert meta["source_format"] == "xlsx"
        assert meta["csv_schema"] == "pv_marked_register_v1"
        assert meta["rows_read"] == 2
        assert meta["postal_rows"] == 2
        assert meta["voted_rows"] == 1

    def test_pv_marked_register_extracts_by_heading_when_columns_move(
        self, tmp_path
    ):
        headers = list(reversed(PV_MARKED_HEADERS))
        source_row = _pv_marked_row(
            elector_number="2B-91",
            status="6",
            receipt_date=date(2026, 7, 9),
        )
        values_by_header = dict(zip(PV_MARKED_HEADERS, source_row))
        reordered_row = [values_by_header[header] for header in headers]
        path = _save_workbook(
            tmp_path,
            _pv_marked_workbook(reordered_row, headers=headers),
        )

        rows, _ = h._parse_uploaded_xlsx(path, "Test", "Date")

        assert rows[0]["polling_district"] == "2B"
        assert rows[0]["elector_number"] == "91"
        assert rows[0]["voted"] == "Y"

    @pytest.mark.parametrize(
        ("row_overrides", "expected_detail", "private_value"),
        [
            (
                {"elector_number": "PRIVATE_BAD_REFERENCE"},
                "rows with invalid ElectorNo: 1",
                "PRIVATE_BAD_REFERENCE",
            ),
            (
                {"status": "PRIVATE_STATUS"},
                "rows with invalid PVSStatus: 1",
                "PRIVATE_STATUS",
            ),
            (
                {"status": "1", "receipt_date": "PRIVATE_RECEIPT"},
                "rows with invalid DecReceiptDate: 1",
                "PRIVATE_RECEIPT",
            ),
            (
                {"status": "0", "receipt_date": "09/07/2026 21:14:05"},
                "rows with inconsistent postal-vote markers: 1",
                "09/07/2026 21:14:05",
            ),
            (
                {"status": "1", "receipt_date": ""},
                "rows with inconsistent postal-vote markers: 1",
                None,
            ),
        ],
    )
    def test_pv_marked_register_rejects_invalid_rows_aggregately(
        self,
        tmp_path,
        row_overrides,
        expected_detail,
        private_value,
    ):
        path = _save_workbook(
            tmp_path,
            _pv_marked_workbook(_pv_marked_row(**row_overrides)),
        )

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_ROW_INVALID"
        assert expected_detail in str(exc.value)
        if private_value:
            assert private_value not in str(exc.value)

    def test_pv_marked_register_rejects_duplicate_references(self, tmp_path):
        path = _save_workbook(
            tmp_path,
            _pv_marked_workbook(
                _pv_marked_row(),
                _pv_marked_row(record_number="2"),
            ),
        )

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_ROW_INVALID"
        assert "duplicate district/elector keys: 1" in str(exc.value)

    def test_pv_marked_register_requires_the_complete_header_profile(
        self, tmp_path
    ):
        headers = list(PV_MARKED_HEADERS)
        headers[-2] = "UnknownStatus"
        path = _save_workbook(
            tmp_path,
            _pv_marked_workbook(_pv_marked_row(), headers=headers),
        )

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_HEADER_UNRECOGNISED",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    def test_selects_one_recognised_visible_sheet_among_notes(self, tmp_path):
        workbook = Workbook()
        notes = workbook.active
        notes.title = "Instructions"
        notes.append(["Synthetic instructions"])
        register = workbook.create_sheet("Register")
        _populate_flat_sheet(register)
        path = _save_workbook(tmp_path, workbook)

        rows, _ = h._parse_uploaded_xlsx(path, "Test", "Date")

        assert len(rows) == 1
        assert rows[0]["polling_district"] == "PD1"

    def test_two_recognised_sheets_are_ambiguous(self, tmp_path):
        workbook = _flat_workbook()
        second = workbook.create_sheet("Second register")
        _populate_flat_sheet(second, district="PD2", elector_number="48")
        path = _save_workbook(tmp_path, workbook)

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_WORKSHEET_AMBIGUOUS",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    def test_hidden_recognised_sheet_is_rejected(self, tmp_path):
        workbook = Workbook()
        notes = workbook.active
        notes.title = "Instructions"
        notes.append(["Synthetic instructions"])
        register = workbook.create_sheet("Register")
        _populate_flat_sheet(register)
        register.sheet_state = "hidden"
        path = _save_workbook(tmp_path, workbook)

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_WORKSHEET_HIDDEN",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    @pytest.mark.parametrize("with_value", [False, True])
    def test_empty_or_unrecognised_workbook_is_rejected(
        self, tmp_path, with_value
    ):
        workbook = Workbook()
        if with_value:
            workbook.active.append(["Unrecognised heading"])
        path = _save_workbook(tmp_path, workbook)

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_HEADER_UNRECOGNISED",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    def test_formula_in_required_data_is_rejected_without_formula_text(
        self, tmp_path
    ):
        formula = "=40+7"
        path = _save_workbook(
            tmp_path,
            _flat_workbook(elector_number=formula),
        )

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_FORMULA_UNSUPPORTED"
        assert formula not in str(exc.value)

    def test_formula_in_discarded_private_column_is_also_rejected(
        self, tmp_path
    ):
        formula = '=HYPERLINK("https://invalid.example","private")'
        path = _save_workbook(
            tmp_path,
            _flat_workbook(private_value=formula),
        )

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_FORMULA_UNSUPPORTED"
        assert formula not in str(exc.value)

    def test_error_cell_is_rejected_without_cell_value(self, tmp_path):
        workbook = _flat_workbook()
        cell = workbook.active.cell(row=2, column=2)
        cell.value = "#VALUE!"
        cell.data_type = "e"
        path = _save_workbook(tmp_path, workbook)

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == "XLSX_CELL_INVALID"
        assert "#VALUE!" not in str(exc.value)

    def test_corrupt_or_non_zip_payload_is_rejected(self, tmp_path):
        path = tmp_path / "invalid.xlsx"
        path.write_bytes(b"not-an-ooxml-archive")

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_ARCHIVE_INVALID",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    def test_macro_member_is_rejected(self, tmp_path):
        path = _save_workbook(tmp_path, _flat_workbook())
        with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("xl/vbaProject.bin", b"synthetic")

        with pytest.raises(
            h.XlsxInputError,
            match="XLSX_ACTIVE_CONTENT_UNSUPPORTED",
        ):
            h._parse_uploaded_xlsx(path, "Test", "Date")

    @pytest.mark.parametrize(
        ("limit_name", "limit_value", "expected_code"),
        [
            ("XLSX_MAX_ARCHIVE_MEMBERS", 1, "XLSX_ARCHIVE_INVALID"),
            ("XLSX_MAX_UNCOMPRESSED_BYTES", 1, "XLSX_ARCHIVE_TOO_LARGE"),
            ("XLSX_MAX_COMPRESSION_RATIO", 1, "XLSX_ARCHIVE_RATIO_INVALID"),
            ("XLSX_MAX_WORKSHEETS", 1, "XLSX_WORKBOOK_INVALID"),
            ("XLSX_MAX_COLUMNS", 2, "XLSX_TOO_MANY_COLUMNS"),
            ("XLSX_MAX_PHYSICAL_ROWS", 1, "XLSX_TOO_MANY_ROWS"),
            ("XLSX_MAX_CELL_CHARS", 5, "XLSX_CELL_TOO_LARGE"),
        ],
    )
    def test_configured_archive_and_sheet_limits_fire_before_extraction(
        self,
        tmp_path,
        monkeypatch,
        limit_name,
        limit_value,
        expected_code,
    ):
        workbook = _flat_workbook()
        if limit_name == "XLSX_MAX_WORKSHEETS":
            workbook.create_sheet("Notes")
        path = _save_workbook(tmp_path, workbook)
        monkeypatch.setattr(h, limit_name, limit_value)

        with pytest.raises(h.XlsxInputError) as exc:
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert exc.value.code == expected_code

    def test_workbook_is_closed_after_a_validation_failure(
        self, tmp_path, monkeypatch
    ):
        formula = "=1+1"
        path = _save_workbook(
            tmp_path,
            _flat_workbook(elector_number=formula),
        )
        real_load_workbook = openpyxl.load_workbook
        closed = []

        def tracked_load_workbook(*args, **kwargs):
            workbook = real_load_workbook(*args, **kwargs)
            real_close = workbook.close

            def tracked_close():
                closed.append(True)
                real_close()

            workbook.close = tracked_close
            return workbook

        monkeypatch.setattr(openpyxl, "load_workbook", tracked_load_workbook)

        with pytest.raises(h.XlsxInputError):
            h._parse_uploaded_xlsx(path, "Test", "Date")

        assert closed == [True]

    def test_openpyxl_uses_defusedxml(self):
        assert openpyxl.DEFUSEDXML is True


class _FakeS3:
    def __init__(self, source_bytes=b"", download_error=None, head_size=None):
        self.source_bytes = source_bytes
        self.download_error = download_error
        self.head_size = head_size
        self.downloads = 0
        self.puts = []

    def download_file(self, _bucket, _key, destination):
        if self.download_error:
            raise self.download_error
        self.downloads += 1
        Path(destination).write_bytes(self.source_bytes)

    def head_object(self, **_kwargs):
        if self.download_error:
            raise self.download_error
        return {
            "ContentLength": (
                self.head_size
                if self.head_size is not None
                else len(self.source_bytes)
            )
        }

    def put_object(self, **kwargs):
        self.puts.append(kwargs)


def _xlsx_job(status="QUEUED"):
    return {
        "jobId": "job-xlsx",
        "fileType": "xlsx",
        "userSub": "user-1",
        "batchId": "batch-1",
        "totalFilesInBatch": 1,
        "association": "Test Association",
        "constituency": "Test Constituency",
        "councilArea": "Test Council",
        "election": "Test Election",
        "electionDate": "01/05/2026",
        "status": status,
    }


def _xlsx_event(receive_count="1"):
    return {
        "Records": [{
            "messageId": "message-xlsx",
            "body": json.dumps({
                "jobId": "job-xlsx",
                "bucket": "input-bucket",
                "s3Key": "uploads/user-1/job-xlsx/input.xlsx",
            }),
            "attributes": {"ApproximateReceiveCount": receive_count},
        }]
    }


class TestXlsxHandlerDispatch:
    def _arrange(self, monkeypatch, fake_s3):
        finalised = []
        monkeypatch.setattr(h, "OCR_AVAILABLE", False)
        monkeypatch.setattr(h, "s3_client", fake_s3)
        monkeypatch.setattr(h, "get_job", lambda _job_id: _xlsx_job())
        monkeypatch.setattr(
            h,
            "_ensure_single_chunk_tracker",
            lambda _job_id: None,
        )
        monkeypatch.setattr(
            h,
            "try_finalise_job",
            lambda *args, **kwargs: finalised.append((args, kwargs)),
        )
        monkeypatch.setattr(
            h,
            "run_splitter",
            lambda *_args, **_kwargs: pytest.fail(
                "XLSX reached the PDF splitter"
            ),
        )
        monkeypatch.setattr(
            h,
            "ocr_pdf",
            lambda *_args, **_kwargs: pytest.fail("XLSX reached OCR"),
        )
        return finalised

    def test_xlsx_bypasses_ocr_and_writes_normalised_json(
        self, monkeypatch
    ):
        fake_s3 = _FakeS3(_workbook_bytes(_flat_workbook()))
        finalised = self._arrange(monkeypatch, fake_s3)

        result = h.handler(_xlsx_event(), None)

        assert result == {"batchItemFailures": []}
        assert len(fake_s3.puts) == 1
        payload = json.loads(fake_s3.puts[0]["Body"])
        assert payload["meta"]["source_type"] == "csv"
        assert payload["meta"]["source_format"] == "xlsx"
        assert payload["rows"][0]["elector_number"] == "47/1"
        assert finalised[0][1]["chunk_failed"] is False

    def test_invalid_xlsx_settles_failed_without_output(self, monkeypatch):
        fake_s3 = _FakeS3(b"not-an-xlsx")
        finalised = self._arrange(monkeypatch, fake_s3)

        result = h.handler(_xlsx_event(), None)

        assert result == {"batchItemFailures": []}
        assert fake_s3.puts == []
        assert finalised[0][1]["chunk_failed"] is True
        assert "XLSX_ARCHIVE_INVALID" in finalised[0][1]["failure_reason"]

    def test_oversize_xlsx_is_rejected_before_download(
        self, monkeypatch
    ):
        fake_s3 = _FakeS3(b"x", head_size=101)
        finalised = self._arrange(monkeypatch, fake_s3)
        monkeypatch.setattr(h, "CSV_MAX_BYTES", 100)

        result = h.handler(_xlsx_event(), None)

        assert result == {"batchItemFailures": []}
        assert fake_s3.downloads == 0
        assert fake_s3.puts == []
        assert finalised[0][1]["chunk_failed"] is True
        assert "XLSX_TOO_LARGE" in finalised[0][1]["failure_reason"]

    def test_xlsx_s3_failure_returns_message_for_retry(self, monkeypatch):
        fake_s3 = _FakeS3(
            download_error=RuntimeError("temporary S3 failure")
        )
        finalised = self._arrange(monkeypatch, fake_s3)

        result = h.handler(_xlsx_event(), None)

        assert result == {
            "batchItemFailures": [{"itemIdentifier": "message-xlsx"}]
        }
        assert finalised == []

    def test_terminal_xlsx_redelivery_remains_idempotent(
        self, monkeypatch
    ):
        fake_s3 = _FakeS3(_workbook_bytes(_flat_workbook()))
        finalised = self._arrange(monkeypatch, fake_s3)
        monkeypatch.setattr(
            h,
            "get_job",
            lambda _job_id: _xlsx_job(status="SUCCEEDED"),
        )

        result = h.handler(_xlsx_event(receive_count="3"), None)

        assert result == {"batchItemFailures": []}
        assert fake_s3.puts == []
        assert finalised[0][1]["chunk_failed"] is False


@pytest.mark.parametrize(
    ("job", "s3_key", "expected"),
    [
        ({"fileType": "XLSX"}, "ignored", "xlsx"),
        ({}, "uploads/input.XLSX", "xlsx"),
        ({"expectedFileType": "xlsx"}, "ignored", "xlsx"),
    ],
)
def test_input_file_type_accepts_xlsx(job, s3_key, expected):
    assert h._input_file_type(job, s3_key) == expected


@pytest.mark.parametrize("suffix", ["xls", "xlsm", "xlsb", "ods"])
def test_input_file_type_rejects_unsupported_workbooks(suffix):
    with pytest.raises(ValueError, match="Unsupported input file type"):
        h._input_file_type({}, f"uploads/input.{suffix}")
